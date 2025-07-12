from datetime import datetime, timedelta
from io import BytesIO
from textwrap import wrap

import logging
import pandas as pd
import random
import requests
import yfinance as yf
from bs4 import BeautifulSoup
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

from .models import UserProfile, Expense, IncomeTracker,ExpenseActivityLog


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import ExpenseActivityLog

@login_required(login_url='/login')
def expense_logs_handler(request):
    user = request.user

    if request.method == 'GET':
        # Fetch logs, corrected to use 'created_at'
        logs = ExpenseActivityLog.objects.filter(user=user).order_by('-created_at')
        logs_list = [
            f"{log.details} — {log.created_at.strftime('%d-%m-%Y %H:%M')}"
            for log in logs
        ]
        return JsonResponse({"status": "success", "logs": logs_list})

    elif request.method == 'POST':
        # Clear logs
        ExpenseActivityLog.objects.filter(user=user).delete()
        return JsonResponse({'status': 'success', 'message': 'All notifications cleared'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})





# import google.generativeai as genai
# from django.conf import settings
# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# import json

# # Gemini setup
# genai.configure(api_key=settings.GOOGLE_API_KEY)
# model = genai.GenerativeModel('gemini-1.5-flash')  # ya 'gemini-2.0-pro' if needed

# @csrf_exempt
# def gemini_chat_view(request):
#     if request.method == 'POST':
#         data = json.loads(request.body)
#         user_message = data.get('message')

#         if not user_message:
#             return JsonResponse({'error': 'No message provided'}, status=400)

#         try:
#             response = model.generate_content(user_message)
#             return JsonResponse({'reply': response.text})
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)
#     return JsonResponse({'error': 'Invalid request method'}, status=405)





# dashboard/views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
import concurrent.futures
from functools import lru_cache

from .nlp.intent_engine import parse_input
from .nlp.response_router import INTENT_TO_FUNCTION
from google.generativeai import GenerativeModel  # already imported in your setup
from django.conf import settings
import google.generativeai as genai

# Gemini setup
genai.configure(api_key=settings.GOOGLE_API_KEY)
model = genai.GenerativeModel('gemini-1.5-flash')  # Or use 'gemini-1.5-pro' if needed

# ✅ Smart Gemini-based input cleaner with timeout + cache
@lru_cache(maxsize=1000)
def clean_input_with_gemini(raw_input):
    prompt = f"""
    Tum ek spelling aur grammar correct karne wale AI ho.
    Niche user ne jo input diya hai usme galti ho sakti hai.
    Tumhe use clean, simple aur correct format me return karna hai — sirf correct version return karo.

    User input: "{raw_input}"
    Cleaned input:
    """

    def call_gemini():
        try:
            response = model.generate_content(prompt)
            return response.text.strip()
        except Exception:
            return raw_input

    try:
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(call_gemini)
            return future.result(timeout=3)  # ⏱️ 3 sec timeout
    except concurrent.futures.TimeoutError:
        return raw_input  # fallback if Gemini is slow
    except Exception:
        return raw_input


@lru_cache(maxsize=1000)
def clean_input_and_explain_with_gemini(raw_input):
    prompt = f"""
    Tum ek smart AI ho. Niche user ka input diya gaya hai jisme spelling ya grammar mistake ho sakti hai.

    Tumhe teen chize return karni hain JSON format me:
    1. cleaned — Input ka correct version (grammar aur spelling sahi ho)
    2. explanation — Short explanation user ke message ka (user-friendly bhasha me)
    3. intent_guess — Tumhara guess ki user ka maksad kya tha (ek word me)

    Format hamesha JSON me do:

    {{
      "cleaned": "<cleaned sentence>",
      "explanation": "<simple explanation>",
      "intent_guess": "<guessed_intent>"
    }}

    User input: "{raw_input}"
    Response:
    """

    def call_gemini():
        try:
            response = model.generate_content(prompt)
            import json
            return json.loads(response.text.strip())
        except Exception:
            return {
                "cleaned": raw_input,
                "explanation": "⚠️ Gemini se explanation nahi mila.",
                "intent_guess": "unknown"
            }

    try:
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(call_gemini)
            return future.result(timeout=5)
    except concurrent.futures.TimeoutError:
        return {
            "cleaned": raw_input,
            "explanation": "⏱️ Gemini slow tha, original input liya.",
            "intent_guess": "unknown"
        }


# ✅ Main chatbot view
@csrf_exempt
def chatbot_view(request):
    if not request.user.is_authenticated:
        return JsonResponse({"response": "🔒 Please login first to use the chatbot."})

    user_input = ""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            user_input = data.get("message", "").strip()
        except json.JSONDecodeError:
            return JsonResponse({"response": "⚠️ Invalid request format."})
    else:
        user_input = request.GET.get("message", "").strip()

    if not user_input:
        return JsonResponse({"response": "❓ Please enter a message to continue."})

    # ✅ Step 1: Gemini clean + explanation
    gemini_result = clean_input_and_explain_with_gemini(user_input)
    cleaned_input = gemini_result["cleaned"]
    ai_explanation = gemini_result["explanation"]
    ai_intent_guess = gemini_result.get("intent_guess", "unknown")

    # ✅ Step 2: NLP intent detection
    intent, entities = parse_input(cleaned_input)

    # ✅ Step 3: App logic call
    if intent != "unknown" and intent in INTENT_TO_FUNCTION:
        try:
            response_func = INTENT_TO_FUNCTION[intent]
            result = response_func(request.user, entities)
            return JsonResponse({
                "intent": intent,
                "entities": entities,
                "input_cleaned": cleaned_input,
                "ai_explanation": ai_explanation,
                "response": result
            })
        except Exception as e:
            return JsonResponse({
                "intent": intent,
                "input_cleaned": cleaned_input,
                "ai_explanation": ai_explanation,
                "response": f"⚠️ Error while handling request: {str(e)}"
            })

    # ❌ Unknown intent
    return JsonResponse({
        "intent": "unknown",
        "input_cleaned": cleaned_input,
        "ai_explanation": ai_explanation,
        "response": "🤖 Sorry, I couldn't understand your message. Please try rephrasing it."
    })



@login_required # This decorator ensures only logged-in users can access this view
def get_user_first_name(request):
    """
    Returns the first name of the logged-in user.
    Requires the user to be authenticated.
    """
    # request.user is available because of @login_required
    # Access first_name directly from the user object
    first_name = request.user.first_name if request.user.first_name else ""
    return JsonResponse({"first_name": first_name})





from datetime import date

@login_required(login_url='/login')
def dashboard_home(request):
    try:
        user_profile = UserProfile.objects.get(user=request.user)

        # Define current calendar year
        current_year = timezone.now().year
        start_date = date(current_year, 1, 1)
        end_date = date(current_year, 12, 31)

        # Total income for the current calendar year
        annual_income = IncomeTracker.objects.filter(
            user=request.user,
            date__range=(start_date, end_date)
        ).aggregate(total=Sum('amount'))['total'] or 0

        # Total expenses for the current calendar year
        total_expense_1y = Expense.objects.filter(
            user=request.user,
            date__range=(start_date, end_date)
        ).aggregate(total=Sum('amount'))['total'] or 0

        context = {
            'user_financial_data': {
                'annual_income': annual_income,
                'savings': user_profile.savings,
                'existing_investments': user_profile.existing_investments,
            },
            'monthly_expense_total': total_expense_1y
        }

        return render(request, 'dashboard.html', context)

    except UserProfile.DoesNotExist:
        return redirect('/create-profile/')



from django.http import JsonResponse
from django.shortcuts import render
import joblib
import pandas as pd
from dashboard.models import StockData
from django.conf import settings
import os

# Function to suggest top stocks (already implemented)
@login_required(login_url='/login')
def suggest_top_stocks_api(request):
    model_dir = os.path.join(settings.BASE_DIR, 'ml', 'models', 'stock')
    timeframes = {
        'next_day': 1,
        '7_days': 7,
        '15_days': 15,
        '1_month': 30,
        '3_months': 90,
        '6_months': 180,
        '1_year': 365
    }

    stock_predictions = []

    for model_file in os.listdir(model_dir):
        if model_file.endswith('_model.pkl'):
            symbol = model_file.split('_')[0]

            try:
                model_path = os.path.join(model_dir, model_file)
                model = joblib.load(model_path)

                latest_stock_data = StockData.objects.filter(symbol=symbol).order_by('-date').first()
                if not latest_stock_data:
                    continue

                latest_price = latest_stock_data.close
                predictions = {}
                for label, days_ahead in timeframes.items():
                    future_df = pd.DataFrame({'days': [days_ahead]})
                    predicted_price = model.predict(future_df)
                    predictions[label] = round(predicted_price[0], 2)

                growth = ((predictions['1_month'] - latest_price) / latest_price) * 100
                stock_predictions.append({
                    'symbol': symbol,
                    'predictions': predictions,
                    'growth_percent': round(growth, 2),
                })

            except Exception as e:
                continue

    stock_predictions.sort(key=lambda x: x['growth_percent'], reverse=True)
    top_5_stocks = stock_predictions[:5]

    return JsonResponse({
        "status": "success",
        "suggested_stocks": top_5_stocks
    })


from django.http import JsonResponse
from django.db.models import Max, Min, F
from dashboard.models import MutualFundData
from datetime import timedelta
@login_required(login_url='/login')
def suggest_top_mutual_funds_api(request):
    # Get all unique schemes
    schemes = MutualFundData.objects.values('scheme_name').distinct()

    fund_growth = []

    for scheme in schemes:
        name = scheme['scheme_name']
        queryset = MutualFundData.objects.filter(scheme_name=name).order_by('date')

        if queryset.count() < 2:
            continue  # Not enough data to compare

        first_nav = queryset.first().net_asset_value
        last_nav = queryset.last().net_asset_value

        if first_nav == 0:
            continue

        growth_percentage = ((last_nav - first_nav) / first_nav) * 100

        fund_growth.append({
            'scheme_name': name,
            'start_nav': first_nav,
            'latest_nav': last_nav,
            'growth_percent': round(growth_percentage, 2)
        })

    # Sort and get top 5
    fund_growth.sort(key=lambda x: x['growth_percent'], reverse=True)
    top_5_funds = fund_growth[:5]

    return JsonResponse({
        "status": "success",
        "suggested_mutual_funds": top_5_funds
    })


from dashboard.models import CryptoHistoryData
from django.conf import settings
import os
import joblib
import pandas as pd
from django.http import JsonResponse

@login_required(login_url='/login')
def suggest_top_cryptos_api(request):
    model_dir = os.path.join(settings.BASE_DIR, 'ml', 'models', 'crypto')
    timeframes = {
        'next_day': 1,
        '7_days': 7,
        '15_days': 15,
        '1_month': 30,
        '3_months': 90,
        '6_months': 180,
        '1_year': 365
    }

    crypto_predictions = []

    for model_file in os.listdir(model_dir):
        if model_file.endswith('_model.pkl'):
            symbol = model_file.split('_')[0]

            try:
                model_path = os.path.join(model_dir, model_file)
                model = joblib.load(model_path)

                # Get the latest data for the cryptocurrency
                latest_data = CryptoHistoryData.objects.filter(symbol=symbol).order_by('-date').first()
                if not latest_data:
                    continue

                latest_price = latest_data.price_inr  # Change to price_inr instead of close
                predictions = {}

                for label, days_ahead in timeframes.items():
                    future_df = pd.DataFrame({'days': [days_ahead]})
                    predicted_price = model.predict(future_df)
                    predictions[label] = round(predicted_price[0], 2)

                # Calculate growth based on the 1-month prediction and the latest price
                growth = ((predictions['1_month'] - latest_price) / latest_price) * 100
                crypto_predictions.append({
                    'symbol': symbol,
                    'predictions': predictions,
                    'growth_percent': round(growth, 2),
                })

            except Exception:
                continue

    # Sort the cryptos based on the growth percentage and select the top 5
    crypto_predictions.sort(key=lambda x: x['growth_percent'], reverse=True)
    top_5 = crypto_predictions[:5]

    return JsonResponse({
        "status": "success",
        "suggested_cryptos": top_5
    })


from dashboard.models import CommodityData
from django.conf import settings
import os
import joblib
import pandas as pd
from django.http import JsonResponse
@login_required(login_url='/login')
def suggest_top_commodities_api(request):
    model_dir = os.path.join(settings.BASE_DIR, 'ml', 'models', 'commodity')
    timeframes = {
        'next_day': 1,
        '7_days': 7,
        '15_days': 15,
        '1_month': 30,
        '3_months': 90,
        '6_months': 180,
        '1_year': 365
    }

    commodity_predictions = []

    # Dynamically get all supported commodities from model files
    supported_commodities = [
        model_file.split('_')[0] for model_file in os.listdir(model_dir) if model_file.endswith('_model.pkl')
    ]

    for model_file in os.listdir(model_dir):
        if model_file.endswith('_model.pkl'):
            # Extract the commodity symbol from the model file name
            symbol = model_file.split('_')[0]  # E.g., "BNO" from "BNO_model.pkl"

            if symbol not in supported_commodities:
                continue

            try:
                model_path = os.path.join(model_dir, model_file)
                model = joblib.load(model_path)

                # Get the latest data for the commodity
                latest_data = CommodityData.objects.filter(name=symbol).order_by('-date').first()
                if not latest_data:
                    continue

                latest_price = latest_data.price_inr  # Price in INR
                predictions = {}

                for label, days_ahead in timeframes.items():
                    future_df = pd.DataFrame({'days': [days_ahead]})
                    predicted_price = model.predict(future_df)
                    predictions[label] = round(predicted_price[0], 2)

                # Calculate growth based on the 1-month prediction and the latest price
                growth = ((predictions['1_month'] - latest_price) / latest_price) * 100
                commodity_predictions.append({
                    'symbol': symbol,
                    'predictions': predictions,
                    'growth_percent': round(growth, 2),
                })

            except Exception as e:
                print(f"Error processing {symbol}: {e}")
                continue

    # Sort the commodities based on the growth percentage and select the top 5
    commodity_predictions.sort(key=lambda x: x['growth_percent'], reverse=True)
    top_5 = commodity_predictions[:5]

    return JsonResponse({
        "status": "success",
        "suggested_commodities": top_5
    })


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from dashboard.models import StockData, CryptoHistoryData, CommodityData, MutualFundData
import os, joblib, pandas as pd
from django.conf import settings

@csrf_exempt
@login_required(login_url='/login')
def compare_assets_api(request):
    asset_type = request.GET.get('asset_type')
    symbols = request.GET.get('symbols', '')
    symbol_list = [s.strip().upper() for s in symbols.split(',') if s.strip()]

    if not asset_type or not symbol_list:
        return JsonResponse({"status": "error", "message": "Invalid input."}, status=400)

    model_dir = os.path.join(settings.BASE_DIR, 'ml', 'models', asset_type)
    timeframes = {
        'next_day': 1,
        '7_days': 7,
        '15_days': 15,
        '1_month': 30,
    }

    results = []

    for symbol in symbol_list:
        try:
            model_path = os.path.join(model_dir, f"{symbol}_model.pkl")
            if not os.path.exists(model_path):
                continue  # Model not found

            model = joblib.load(model_path)
            future_df = pd.DataFrame({'days': list(timeframes.values())})
            preds = model.predict(future_df)
            predictions = {label: round(pred, 2) for label, pred in zip(timeframes.keys(), preds)}

            if asset_type == 'stock':
                latest_data = StockData.objects.filter(symbol=symbol).order_by('-date').first()
                price = latest_data.close if latest_data else None
            elif asset_type == 'crypto':
                latest_data = CryptoHistoryData.objects.filter(symbol=symbol).order_by('-date').first()
                price = latest_data.price_inr if latest_data else None
            elif asset_type == 'commodity':
                latest_data = CommodityData.objects.filter(name=symbol).order_by('-date').first()
                price = latest_data.price_inr if latest_data else None
            elif asset_type == 'mutual_fund':
                data = MutualFundData.objects.filter(scheme_name=symbol).order_by('date')
                if data.exists():
                    price = data.last().net_asset_value
                    first_nav = data.first().net_asset_value
                    growth_percent = ((price - first_nav) / first_nav) * 100 if first_nav else 0
                    predictions = {"growth_percent": round(growth_percent, 2)}
                else:
                    continue
            else:
                continue

            results.append({
                "symbol": symbol,
                "latest_price": round(price, 2) if price else None,
                "predictions": predictions
            })
        except Exception as e:
            print(f"Error comparing {symbol}: {e}")
            continue

    return JsonResponse({
        "status": "success",
        "results": results
    })

from datetime import date
from dateutil.relativedelta import relativedelta
from collections import defaultdict
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Sum
from .models import IncomeTracker

@login_required(login_url='/login')
def income_analysis_view(request):
    user = request.user
    today = timezone.now().date()
    selected_year = int(request.GET.get('year', today.year))

    # Date range for selected year
    start_date = date(selected_year, 1, 1)
    end_date = date(selected_year, 12, 31)

    income_all = IncomeTracker.objects.filter(user=user, date__range=(start_date, end_date))
    total_income = income_all.aggregate(total=Sum('amount'))['total'] or 0

    source_totals = income_all.values('source').annotate(total=Sum('amount')).order_by('-total')
    highest_source = source_totals[0] if source_totals else None
    lowest_source = source_totals.last() if source_totals else None

    # Limit months to current month if current year, else full 12
    current_month_limit = today.month if selected_year == today.year else 12

    # Average monthly income only up to current month
    average_monthly_income = total_income / current_month_limit if total_income else 0

    # Passive income ratio
    passive_sources = ['Investment', 'Gift']
    passive_income_total = income_all.filter(source__in=passive_sources).aggregate(total=Sum('amount'))['total'] or 0
    passive_ratio = (passive_income_total / total_income) * 100 if total_income else 0

    # Monthly entry frequency
    monthly_entry_counts = defaultdict(int)
    for income in income_all:
        if income.date.month <= current_month_limit or selected_year != today.year:
            month = income.date.strftime("%Y-%m")
            monthly_entry_counts[month] += 1

    avg_entry_frequency = sum(monthly_entry_counts.values()) / current_month_limit if current_month_limit else 0

    # Chart data
    source_labels = [item['source'] for item in source_totals]
    source_amounts = [float(item['total']) for item in source_totals]

    # Monthly bar chart labels: only till current month (if current year)
    bar_labels = []
    bar_values = []
    for i in range(1, current_month_limit + 1):
        month_date = date(selected_year, i, 1)
        key = month_date.strftime("%Y-%m")
        bar_labels.append(key)
        bar_values.append(monthly_entry_counts.get(key, 0))

    return JsonResponse({
        'total_income': round(total_income, 2),
        'highest_source': highest_source,
        'lowest_source': lowest_source,
        'average_monthly_income': round(average_monthly_income, 2),
        'passive_ratio': round(passive_ratio, 2),
        'avg_entry_frequency': round(avg_entry_frequency, 1),
        'source_labels': source_labels,
        'source_amounts': source_amounts,
        'bar_labels': bar_labels,
        'bar_values': bar_values
    })





from django.db.models import Sum, Count
from django.utils import timezone
from django.http import JsonResponse
from datetime import datetime
from .models import Expense, BudgetLimit
from django.contrib.auth.decorators import login_required
import calendar

@login_required(login_url='/login')
def expense_analysis_view(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'User not authenticated'}, status=401)

    user = request.user
    current_date = timezone.now()
    year = int(request.GET.get('year', current_date.year))

    expenses = Expense.objects.filter(user=user, date__year=year)
    total_expense = expenses.aggregate(total=Sum('amount'))['total'] or 0

    current_month_limit = current_date.month if year == current_date.year else 12
    monthly_data = {}
    for month in range(1, current_month_limit + 1):
        month_label = f"{year}-{month:02d}"
        monthly_exp = expenses.filter(date__month=month).aggregate(total=Sum('amount'))['total'] or 0
        monthly_data[month_label] = float(monthly_exp)

    bar_labels = list(monthly_data.keys())
    bar_values = list(monthly_data.values())

    monthly_growth = []
    for i in range(1, len(bar_values)):
        prev = bar_values[i - 1] or 1
        growth = ((bar_values[i] - prev) / prev) * 100
        monthly_growth.append(round(growth, 2))

    category_data = expenses.values('category').annotate(total=Sum('amount')).order_by('-total')
    category_list = list(category_data)

    category_labels = [entry['category'] for entry in category_list]
    category_totals = [float(entry['total']) for entry in category_list]
    category_percentages = [(entry['total'] / total_expense) * 100 if total_expense else 0 for entry in category_list]

    budget_limits = BudgetLimit.objects.filter(user=user, frequency='monthly')
    budget_comparison = []
    for limit in budget_limits:
        actual = expenses.filter(category=limit.category).aggregate(total=Sum('amount'))['total'] or 0
        budget_comparison.append({
            'category': limit.category,
            'limit': float(limit.limit),
            'actual': float(actual),
            'percentage_used': round((actual / limit.limit) * 100, 2) if limit.limit else 0,
        })

    highest = category_list[0] if category_list else {'category': '—', 'total': 0}
    lowest = category_list[-1] if len(category_list) > 1 else highest

    entries_per_month = expenses.extra({'month': "strftime('%%Y-%%m', date)"}).values('month').annotate(count=Count('id'))
    frequency_map = {entry['month']: entry['count'] for entry in entries_per_month}
    avg_frequency = sum(frequency_map.values()) / len(frequency_map) if frequency_map else 0

    weekday_data = [0] * 7
    for exp in expenses:
        weekday = exp.date.weekday()
        weekday_data[weekday] += float(exp.amount)
    weekday_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    today = timezone.now().date()
    current_day = today.day if today.year == year else 31
    dom_data = [0] * current_day
    dom_labels = list(range(1, current_day + 1))
    current_month_expenses = expenses.filter(date__month=today.month) if today.year == year else Expense.objects.none()
    for exp in current_month_expenses:
        day = exp.date.day
        if day <= current_day:
            dom_data[day - 1] += float(exp.amount)

    return JsonResponse({
        'total_expense': float(total_expense),
        'bar_labels': bar_labels,
        'bar_values': bar_values,
        'monthly_growth': monthly_growth,
        'category_labels': category_labels,
        'category_totals': category_totals,
        'category_percentages': category_percentages,
        'budget_comparison': budget_comparison,
        'highest_category': {
            'category': highest['category'],
            'total': float(highest['total']),
        },
        'lowest_category': {
            'category': lowest['category'],
            'total': float(lowest['total']),
        },
        'avg_entry_frequency': round(avg_frequency, 2),
        'weekday_labels': weekday_labels,
        'weekday_data': weekday_data,
        'dom_labels': dom_labels,
        'dom_data': dom_data,
    })




from django.http import JsonResponse
from django.db.models import Sum
from django.utils.timezone import now
from .models import IncomeTracker, Expense, BudgetLimit
import calendar
from django.contrib.auth.decorators import login_required

@login_required(login_url='/login')
def combined_analysis_view(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'User not authenticated'}, status=401)

    user = request.user
    today = now().date()
    year = int(request.GET.get('year', today.year))
    current_month_limit = today.month if year == today.year else 12

    total_income = IncomeTracker.objects.filter(user=user, date__year=year).aggregate(total=Sum('amount'))['total'] or 0
    total_expense = Expense.objects.filter(user=user, date__year=year).aggregate(total=Sum('amount'))['total'] or 0
    income_expense_ratio = round(total_income / total_expense, 2) if total_expense > 0 else '∞'

    monthly_flow = {}
    for month in range(1, current_month_limit + 1):
        month_name = f"{calendar.month_abbr[month]} {str(year)[-2:]}"  # Example: Jan 25

        income_sum = IncomeTracker.objects.filter(user=user, date__year=year, date__month=month).aggregate(total=Sum('amount'))['total'] or 0
        expense_sum = Expense.objects.filter(user=user, date__year=year, date__month=month).aggregate(total=Sum('amount'))['total'] or 0

        monthly_flow[month_name] = {
            "income": float(income_sum),
            "expense": float(expense_sum),
        }

    category_analysis = []
    categories = [choice[0] for choice in Expense.CATEGORY_CHOICES]
    for cat in categories:
        budget = BudgetLimit.objects.filter(user=user, category=cat).aggregate(total=Sum('limit'))['total'] or 0
        actual = Expense.objects.filter(user=user, category=cat, date__year=year).aggregate(total=Sum('amount'))['total'] or 0
        remaining = max(0, budget - actual)
        category_analysis.append({
            "category": cat,
            "budget": float(budget),
            "actual": float(actual),
            "remaining": float(remaining),
        })

    accuracy_scores = []
    for item in category_analysis:
        budget = item["budget"]
        actual = item["actual"]
        if budget > 0:
            score = max(0, 100 - abs((actual - budget) / budget) * 100)
            accuracy_scores.append(score)
    avg_accuracy = round(sum(accuracy_scores) / len(accuracy_scores), 2) if accuracy_scores else None

    ideal_ratio = {"spend": 50, "save": 30, "invest": 20}
    if total_income > 0:
        actual_spend_percent = round((total_expense / total_income) * 100, 2)
        actual_save_percent = round(100 - actual_spend_percent, 2)
        actual_invest_percent = 0
    else:
        actual_spend_percent = actual_save_percent = actual_invest_percent = 0

    actual_ratio = {
        "spend": actual_spend_percent,
        "save": actual_save_percent,
        "invest": actual_invest_percent,
    }

    data = {
        "total_income": float(total_income),
        "total_expense": float(total_expense),
        "income_expense_ratio": income_expense_ratio,
        "budget_accuracy_score": avg_accuracy,
        "ideal_vs_actual_ratio": {
            "ideal": ideal_ratio,
            "actual": actual_ratio
        },
        "budget_vs_actual": category_analysis,
        "monthly_cash_flow": monthly_flow,
    }

    return JsonResponse(data)







# ✅ Expemse tracker views

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Expense, BudgetLimit, UserProfile
from .forms import ExpenseForm
from django.utils import timezone
from datetime import date, timedelta
from django.db.models.functions import TruncMonth
from django.db.models import Sum
import calendar

@login_required(login_url='/login')
def expense_tracker(request):
    user = request.user
    category_filter = request.GET.get('category', '')
    date_filter = request.GET.get('date', '')
    selected_year = int(request.GET.get('year', timezone.now().year))

    start_of_year = date(selected_year, 1, 1)
    end_of_year = date(selected_year, 12, 31)

    all_expenses = Expense.objects.filter(user=user, date__range=(start_of_year, end_of_year))
    filtered_expenses = all_expenses

    if category_filter:
        filtered_expenses = filtered_expenses.filter(category=category_filter)

    if date_filter:
        filtered_expenses = filtered_expenses.filter(date=date_filter)

    total_expenses = sum(exp.amount for exp in filtered_expenses)
    expenses = filtered_expenses.order_by('-date', '-time')[:10]

    current_month = timezone.now().month if selected_year == timezone.now().year else 12
    monthly_expenses = all_expenses.filter(date__month=current_month)
    monthly_expense_total = sum(exp.amount for exp in monthly_expenses)

    user_financial_data = UserProfile.objects.get(user=user)

    monthly_data = (
        filtered_expenses.annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )

    months = []
    monthly_totals = []
    for entry in monthly_data:
        month_num = entry['month'].month
        months.append(calendar.month_name[month_num])
        monthly_totals.append(float(entry['total']))

    category_data = (
        filtered_expenses.values('category')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )

    overall_total = sum(exp.amount for exp in all_expenses) or 1
    categories = []
    category_totals = []
    category_percentages = []

    for item in category_data:
        categories.append(item['category'])
        category_totals.append(float(item['total']))
        percentage = (item['total'] / overall_total) * 100
        category_percentages.append(round(percentage, 2))

    # Budget Summary Data (still calculated for last 30 and 365 days from today)
    today = timezone.now().date()
    one_month_ago = today - timedelta(days=30)
    one_year_ago = today - timedelta(days=365)

    budget_categories = [choice[0] for choice in Expense.CATEGORY_CHOICES]
    budget_data = []

    for category in budget_categories:
        monthly_limit_obj = BudgetLimit.objects.filter(user=user, category=category, frequency='monthly').first()
        monthly_limit = monthly_limit_obj.limit if monthly_limit_obj else 0

        annual_limit_obj = BudgetLimit.objects.filter(user=user, category=category, frequency='annually').first()
        annual_limit = annual_limit_obj.limit if annual_limit_obj else 0

        monthly_spent = Expense.objects.filter(user=user, category=category, date__gte=one_month_ago).aggregate(total=Sum('amount'))['total'] or 0
        annual_spent = Expense.objects.filter(user=user, category=category, date__gte=one_year_ago).aggregate(total=Sum('amount'))['total'] or 0

        budget_data.append({
            'category': category,
            'monthly_limit': monthly_limit,
            'annual_limit': annual_limit,
            'monthly_spent': monthly_spent,
            'monthly_remaining': max(monthly_limit - monthly_spent, 0),
            'annual_spent': annual_spent,
            'annual_remaining': max(annual_limit - annual_spent, 0),
        })

    context = {
        'expenses': expenses,
        'total_expenses': total_expenses,
        'monthly_expense_total': monthly_expense_total,
        'user_financial_data': user_financial_data,
        'category_filter': category_filter,
        'date_filter': date_filter,
        'months': months,
        'monthly_totals': monthly_totals,
        'categories': categories,
        'category_totals': category_totals,
        'category_percentages': category_percentages,
        'budget_data': budget_data,
        'selected_year': selected_year,
    }

    return render(request, 'expense_tracker.html', context)



from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404

@login_required(login_url='/login')
@require_POST
def delete_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    expense.delete()
    return JsonResponse({'success': True})




from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils import timezone
from django.http import JsonResponse
from django.contrib import messages
from django.utils.formats import number_format
from datetime import date, timedelta

from .models import Expense
from .forms import ExpenseForm


@login_required(login_url='/login')
def add_expense_view(request):
    today = timezone.now().date()
    current_year = today.year
    start_of_year = date(current_year, 1, 1)
    end_of_year = date(current_year, 12, 31)
    thirty_days_ago = today - timedelta(days=30)

    def get_category_summary():
        today = timezone.now().date()
        current_year = today.year
        current_month = today.month
        start_of_year  = date(current_year, 1, 1)
        start_of_month = date(current_year, current_month, 1)
    
        total_qs = Expense.objects.filter(
            user=request.user,
            date__range=(start_of_year, today)       # entire year up to today
        ).values('category').annotate(total=Sum('amount')).order_by('category')
    
        # THIS MONTH: from the 1st to today
        recent_qs = Expense.objects.filter(
            user=request.user,
            date__range=(start_of_month, today)
        ).values('category').annotate(total=Sum('amount')).order_by('category')
    
        category_map = dict(Expense.CATEGORY_CHOICES)
        summary = {}
        raw_total = raw_recent = 0
    
        for item in total_qs:
            cat = category_map[item['category']]
            amt = item['total'] or 0
            summary[cat] = {'total': amt, 'recent': 0}
            raw_total += amt
    
        for item in recent_qs:
            cat = category_map[item['category']]
            amt = item['total'] or 0
            summary.setdefault(cat, {'total': 0, 'recent': 0})
            summary[cat]['recent'] = amt
            raw_recent += amt
    
        # Format with commas
        for cat, vals in summary.items():
            vals['total']  = number_format(vals['total'],  use_l10n=True)
            vals['recent'] = number_format(vals['recent'], use_l10n=True)
    
        total_recent_sum  = number_format(raw_recent, use_l10n=True)
        total_expense_sum = number_format(raw_total,  use_l10n=True)
    
        return summary, total_recent_sum, total_expense_sum


    form = ExpenseForm(request.POST or None)
    category_expenses, total_recent_sum, total_expense_sum = get_category_summary()

    if request.method == 'POST':
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()

            # re-fetch after save
            category_expenses, total_recent_sum, total_expense_sum = get_category_summary()
            formatted_amt = number_format(expense.amount, use_l10n=True)
            cat_name = dict(Expense.CATEGORY_CHOICES).get(expense.category, expense.category)
            success_msg = f"Expense of ₹{formatted_amt} added under '{cat_name}' category!"

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_msg,
                    'category_expenses': category_expenses,
                    'total_recent_sum': total_recent_sum,
                    'total_expense_sum': total_expense_sum,
                })

            messages.success(request, success_msg)
            return render(request, 'add_expense.html', {
                'form': ExpenseForm(),
                'category_expenses': category_expenses,
                'total_recent_sum': total_recent_sum,
                'total_expense_sum': total_expense_sum,
            })
        else:
            error_msg = "Form is invalid. Please check your input."
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})

            messages.error(request, error_msg)

    return render(request, 'add_expense.html', {
        'form': form,
        'category_expenses': category_expenses,
        'total_recent_sum': total_recent_sum,
        'total_expense_sum': total_expense_sum,
    })



from decimal import Decimal, InvalidOperation
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib import messages
from django.http import JsonResponse
from .models import BudgetLimit, Expense

@login_required(login_url='/login')
def manage_budget_limits(request):
    category_choices = Expense.CATEGORY_CHOICES
    frequency_choices = dict(BudgetLimit.FREQUENCY_CHOICES)

    budget_limits_qs = BudgetLimit.objects.filter(user=request.user).order_by('category')

    if request.method == 'POST':
        action = request.POST.get('action', '').lower()
        category = request.POST.get('category')
        frequency = request.POST.get('frequency', 'monthly')

        if not category:
            error_message = 'Category is required.'
            return _handle_error(request, error_message, budget_limits_qs, category_choices)

        if frequency not in frequency_choices:
            error_message = 'Invalid frequency selected.'
            return _handle_error(request, error_message, budget_limits_qs, category_choices)

        if action == 'delete':
            BudgetLimit.objects.filter(user=request.user, category=category, frequency=frequency).delete()
            success_message = f'Budget limit for {dict(category_choices).get(category)} ({frequency}) deleted successfully.'
            return _handle_success(request, success_message, budget_limits_qs, category_choices, frequency_choices)

        limit = request.POST.get('limit')
        if not limit:
            error_message = 'Please enter a limit.'
            return _handle_error(request, error_message, budget_limits_qs, category_choices)

        try:
            limit_decimal = Decimal(limit)
            if limit_decimal <= 0:
                raise InvalidOperation

            existing_limit = BudgetLimit.objects.filter(user=request.user, category=category, frequency=frequency).first()
            if existing_limit:
                existing_limit.limit = limit_decimal
                existing_limit.save()
                success_message = f'Budget limit for {existing_limit.get_category_display()} ({existing_limit.get_frequency_display()}) updated successfully.'
            else:
                BudgetLimit.objects.create(user=request.user, category=category, limit=limit_decimal, frequency=frequency)
                success_message = f'Budget limit for {dict(category_choices).get(category)} ({frequency_choices.get(frequency)}) set successfully.'

            return _handle_success(request, success_message, budget_limits_qs, category_choices, frequency_choices)

        except InvalidOperation:
            error_message = 'Invalid limit value.'
            return _handle_error(request, error_message, budget_limits_qs, category_choices)

    # GET request
    return render(request, 'manage_budget_limits.html', {
        'budget_limits': budget_limits_qs,
        'category_choices': category_choices,
        'frequency_choices': BudgetLimit.FREQUENCY_CHOICES,
    })


def _handle_error(request, error_message, budget_limits_qs, category_choices):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': error_message})
    messages.error(request, error_message)
    return render(request, 'manage_budget_limits.html', {
        'budget_limits': budget_limits_qs,
        'category_choices': category_choices,
        'frequency_choices': BudgetLimit.FREQUENCY_CHOICES,
    })


def _handle_success(request, success_message, budget_limits_qs, category_choices, frequency_choices):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        updated_limits = budget_limits_qs.values('category', 'limit', 'created_at', 'frequency')
        updated_limits_list = [{
            'category': item['category'],
            'category_display': dict(category_choices).get(item['category']),
            'limit': str(item['limit']),
            'created_at': item['created_at'].isoformat(),
            'frequency': item['frequency'],
            'frequency_display': frequency_choices.get(item['frequency']),
        } for item in updated_limits]
        return JsonResponse({'success': True, 'message': success_message, 'budget_limits': updated_limits_list})

    messages.success(request, success_message)
    return render(request, 'manage_budget_limits.html', {
        'budget_limits': budget_limits_qs,
        'category_choices': category_choices,
        'frequency_choices': BudgetLimit.FREQUENCY_CHOICES,
    })



from django.http import HttpResponse
from .models import Expense
from django.contrib.auth.decorators import login_required
import csv
import openpyxl
from io import BytesIO
from datetime import datetime

@login_required(login_url='/login')
def download_expenses(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    file_format = request.GET.get('format')

    if not start_date or not end_date or not file_format:
        return HttpResponse("Missing data", status=400)

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Filter expenses
    expenses = Expense.objects.filter(
        user=request.user,
        date__gte=start,
        date__lte=end
    )

    if file_format == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="expenses_{start_date}_to_{end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Date', 'Time', 'Category', 'Amount (INR)', 'Description'])

        for exp in expenses:
            writer.writerow([
                exp.date.strftime("%Y-%m-%d"),
                exp.time.strftime("%H:%M:%S"),
                exp.category,
                exp.amount,
                exp.description
            ])

        return response

    elif file_format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(['Date', 'Time', 'Category', 'Amount (INR)', 'Description'])

        for exp in expenses:
            ws.append([
                exp.date.strftime("%Y-%m-%d"),
                exp.time.strftime("%H:%M:%S"),
                exp.category,
                float(exp.amount),
                exp.description
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="expenses_{start_date}_to_{end_date}.xlsx"'
        return response

    else:
        return HttpResponse("Invalid format", status=400)





# income Tracker views
from datetime import date, datetime
from dateutil.relativedelta import relativedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.utils.timezone import make_aware

from .forms import IncomeTrackerForm
from .models import IncomeTracker, UserProfile
@login_required(login_url='/login')
def income_tracker(request):
    source_filter = request.GET.get('source', '')
    date_filter = request.GET.get('date', '')

    today = timezone.now().date()
    current_year = today.year
    current_month = today.month

    start_of_year = date(current_year, 1, 1)
    end_of_year = date(current_year, 12, 31)

    # Filter incomes for full current year: Jan 1 to Dec 31
    all_incomes = IncomeTracker.objects.filter(user=request.user, date__gte=start_of_year, date__lte=end_of_year)
    filtered_incomes = all_incomes

    if source_filter:
        filtered_incomes = filtered_incomes.filter(source=source_filter)

    if date_filter:
        filtered_incomes = filtered_incomes.filter(date=date_filter)

    total_incomes = sum(income.amount for income in filtered_incomes)
    incomes = filtered_incomes.order_by('-date', '-time')[:10]

    # Current month income total from filtered data
    monthly_incomes = all_incomes.filter(date__month=current_month, date__year=current_year)
    monthly_income_total = sum(income.amount for income in monthly_incomes)

    user_financial_data = UserProfile.objects.get(user=request.user)

    # Monthly chart data: group by month and sum amounts (within current year only)
    monthly_data = (
        filtered_incomes.annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )

    # Prepare months list and totals from Jan to current month only
    months = []
    monthly_totals = []
    month_index_map = {}

    for m in range(1, current_month + 1):
        dt = date(current_year, m, 1)
        label = dt.strftime("%b-%Y")  # Eg: Jan-2025
        months.append(label)
        monthly_totals.append(0)
        month_index_map[dt.strftime("%Y-%m")] = m - 1

    for entry in monthly_data:
        key = entry['month'].strftime("%Y-%m")
        if key in month_index_map:
            idx = month_index_map[key]
            monthly_totals[idx] = float(entry['total'])

    # Source pie chart data
    source_data = (
        filtered_incomes.values('source')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )

    overall_total = total_incomes or 1  # prevent division by zero
    sources = []
    source_totals = []
    source_percentages = []

    for item in source_data:
        sources.append(item['source'])
        source_totals.append(float(item['total']))
        percentage = (item['total'] / overall_total) * 100
        source_percentages.append(round(percentage, 2))

    context = {
        'incomes': incomes,
        'total_incomes': total_incomes,
        'monthly_income_total': monthly_income_total,
        'user_financial_data': user_financial_data,
        'source_filter': source_filter,
        'months': months,
        'monthly_totals': monthly_totals,
        'sources': sources,
        'source_totals': source_totals,
        'source_percentages': source_percentages,
        'date_filter': date_filter,
    }

    return render(request, 'income_tracker.html', context)





from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils import timezone
from django.http import JsonResponse
from django.contrib import messages
from datetime import date
from .models import IncomeTracker
from .forms import IncomeTrackerForm
from django.utils.formats import number_format

@login_required(login_url='/login')
def add_income_view(request):
    today = timezone.now().date()
    current_year = today.year
    current_month = today.month
    start_of_year = date(current_year, 1, 1)
    end_of_year = date(current_year, 12, 31)
    start_of_month = date(current_year, current_month, 1)

    def get_category_summary():
        total_income = IncomeTracker.objects.filter(
            user=request.user, date__range=(start_of_year, end_of_year)
        ).values('source').annotate(total=Sum('amount')).order_by('source')
    
        recent_income = IncomeTracker.objects.filter(
            user=request.user, date__range=(start_of_month, today)
        ).values('source').annotate(total=Sum('amount')).order_by('source')
    
        category_income = {}
        source_map = dict(IncomeTracker.SOURCE_CHOICES)
    
        raw_total_sum = 0
        raw_recent_sum = 0
    
        # Add total values first
        for item in total_income:
            source = source_map.get(item['source'], item['source'])
            total = item['total'] or 0
            category_income[source] = {'total': total, 'recent': 0}
            raw_total_sum += total
    
        # Add recent values
        for item in recent_income:
            source = source_map.get(item['source'], item['source'])
            recent = item['total'] or 0
            if source in category_income:
                category_income[source]['recent'] = recent
            else:
                category_income[source] = {'total': 0, 'recent': recent}
            raw_recent_sum += recent
    
        # Format values with commas
        for source in category_income:
            category_income[source]['total'] = number_format(category_income[source]['total'], use_l10n=True)
            category_income[source]['recent'] = number_format(category_income[source]['recent'], use_l10n=True)
    
        total_recent_sum = number_format(raw_recent_sum, use_l10n=True)
        total_income_sum = number_format(raw_total_sum, use_l10n=True)
    
        return category_income, total_recent_sum, total_income_sum

    form = IncomeTrackerForm(request.POST or None)
    category_income, total_recent_sum, total_income_sum = get_category_summary()

    if request.method == 'POST':
        if form.is_valid():
            income = form.save(commit=False)
            income.user = request.user
            income.save()

            # ✅ success message
            source_map = dict(IncomeTracker.SOURCE_CHOICES)
            category_name = source_map.get(income.source, income.source)
            formatted_amount = number_format(income.amount, use_l10n=True)
            success_msg = f"Income of ₹{formatted_amount} added under '{category_name}' category!"

            return handle_income_success(
                request, success_msg,
                IncomeTrackerForm(), category_income, total_recent_sum, total_income_sum
            )
        else:
            return handle_income_error(
                request, "Form is invalid. Please check your input.",
                form, category_income, total_recent_sum, total_income_sum
            )

    return render(request, 'add_income.html', {
        'form': form,
        'category_income': category_income,
        'total_recent_sum': total_recent_sum,
        'total_income_sum': total_income_sum,
    })


def handle_income_error(request, error_message, form, category_income, total_recent_sum, total_income_sum):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': error_message})

    messages.error(request, error_message)
    return render(request, 'add_income.html', {
        'form': form,
        'category_income': category_income,
        'total_recent_sum': total_recent_sum,
        'total_income_sum': total_income_sum,
    })


def handle_income_success(request, success_message, form, category_income, total_recent_sum, total_income_sum):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': success_message})

    messages.success(request, success_message)
    return render(request, 'add_income.html', {
        'form': form,
        'category_income': category_income,
        'total_recent_sum': total_recent_sum,
        'total_income_sum': total_income_sum,
    })






from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from .models import IncomeTracker  # Make sure this import is correct

@login_required(login_url='/login')
@require_POST
def delete_income(request, income_id):
    income = get_object_or_404(IncomeTracker, id=income_id, user=request.user)
    income.delete()
    return JsonResponse({'success': True})



from django.http import HttpResponse
from .models import IncomeTracker
from django.contrib.auth.decorators import login_required
import csv
import openpyxl
from io import BytesIO
from datetime import datetime

@login_required(login_url='/login')
def download_income(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    file_format = request.GET.get('format')

    if not start_date or not end_date or not file_format:
        return HttpResponse("Missing data", status=400)

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Filter income
    incomes = IncomeTracker.objects.filter(
        user=request.user,
        date__gte=start,
        date__lte=end
    )

    if file_format == "csv":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="income_{start_date}_to_{end_date}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Date', 'Time', 'Source', 'Amount (INR)', 'Description'])

        for income in incomes:
            writer.writerow([
                income.date.strftime("%Y-%m-%d"),
                income.time.strftime("%H:%M:%S"),
                income.source,
                income.amount,
                income.description
            ])

        return response

    elif file_format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income"
        ws.append(['Date', 'Time', 'Source', 'Amount (INR)', 'Description'])

        for income in incomes:
            ws.append([
                income.date.strftime("%Y-%m-%d"),
                income.time.strftime("%H:%M:%S"),
                income.source,
                float(income.amount),
                income.description
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="income_{start_date}_to_{end_date}.xlsx"'
        return response

    else:
        return HttpResponse("Invalid format", status=400)








@login_required(login_url='/login')
def market_update(request):
    return render(request, 'market_update.html')



# ✅ Market Section
# ✅ Market Indices
@login_required(login_url='/login')
def get_market_indices(request):
    indices = {
        "Nifty 50": "^NSEI",
        "Sensex": "^BSESN",
        "Bank Nifty": "^NSEBANK",
        "Nifty IT": "^CNXIT",
        "Nifty Midcap 100": "^NSEMDCP50",
    }

    market_data = {}
    for index, symbol in indices.items():
        try:
            data = yf.Ticker(symbol)
            hist = data.history(period="2d")  # Fetching last 2 days of data

            if len(hist) >= 2:
                latest_price = round(float(hist["Close"].iloc[-1]), 2)
                prev_price = round(float(hist["Close"].iloc[-2]), 2)
                change = round(latest_price - prev_price, 2)
                percent = round((change / prev_price) * 100, 2)
            else:
                latest_price, change, percent = "N/A", 0, 0

            market_data[index] = {
                "price": latest_price,
                "change": change,
                "percent": percent,
            }

        except Exception as e:
            print(f"❌ Error fetching {index}: {e}")  # Log error
            market_data[index] = {"price": "N/A", "change": 0, "percent": 0}

    return JsonResponse(market_data)





from django.views.decorators.http import require_GET
@login_required(login_url='/login')
@require_GET
def get_index_chart(request, index_name):
    symbols = {
        "Nifty 50": "^NSEI",
        "Sensex": "^BSESN",
        "Bank Nifty": "^NSEBANK",
        "Nifty IT": "^CNXIT",
        "Nifty Midcap 100": "^NSEMDCP50",
    }

    symbol = symbols.get(index_name)
    if not symbol:
        return JsonResponse({"error": "Invalid index name."})

    period = request.GET.get("period", "1mo")
    allowed_periods = ["1d", "1wk", "1mo", "3mo", "6mo", "1y", "2y", "5y", "10y", "ytd", "max"]
    if period not in allowed_periods:
        return JsonResponse({"error": "Invalid period."})

    try:
        ticker = yf.Ticker(symbol)

        # ⏱️ Use 5-minute interval for 1D to show full day movement
        if period == "1d":
            hist = ticker.history(period="1d", interval="5m")
        else:
            hist = ticker.history(period=period)

        dates = [dt.strftime("%Y-%m-%d %H:%M") if period == "1d" else str(dt.date()) for dt in hist.index]
        prices = [round(price, 2) for price in hist["Close"]]

        return JsonResponse({"dates": dates, "prices": prices})
    except Exception as e:
        return JsonResponse({"error": str(e)})



# ✅ Currency & Commodities
# 🟢 **Currency & Commodities Fetch Function**
@login_required(login_url='/login')
def get_currency_commodities(request):
    try:
        # Helper to calculate percent change
        def get_price_and_change(ticker):
            data = yf.Ticker(ticker).history(period="2d")['Close']
            if len(data) < 2:
                return round(data[-1], 2), 0.00
            current = data.iloc[-1]
            previous = data.iloc[-2]
            change = ((current - previous) / previous) * 100
            return round(current, 2), round(change, 2)

        # Commodities
        gold_price, gold_change = get_price_and_change("GC=F")
        silver_price, silver_change = get_price_and_change("SI=F")
        platinum_price, platinum_change = get_price_and_change("PL=F")

        # Currencies
        usd_inr, usd_inr_change = get_price_and_change("USDINR=X")
        eur_inr, eur_inr_change = get_price_and_change("EURINR=X")

        data = {
            "gold": gold_price,
            "gold_change": gold_change,
            "silver": silver_price,
            "silver_change": silver_change,
            "platinum": platinum_price,
            "platinum_change": platinum_change,
            "usd_inr": usd_inr,
            "usd_inr_change": usd_inr_change,
            "eur_inr": eur_inr,
            "eur_inr_change": eur_inr_change
        }

        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ✅ Stock Highlights
# 🟢 **Stock Highlights Fetch Function**
import yfinance as yf
import pandas as pd
import logging
from datetime import datetime
from django.http import JsonResponse

# Logging setup
logger = logging.getLogger(__name__)

# Load Nifty 100 stock symbols from CSV (only once at module load)
try:
    df = pd.read_csv('dashboard/static/csv/nifty_50_stocks.csv')

    nifty_100_stocks = df['Symbol'].tolist()
except Exception as e:
    logger.error(f"Error loading CSV: {e}")
    nifty_100_stocks = []

@login_required(login_url='/login')
def get_stock_highlights(request):
    data = {
        "most_active": [],
        "fifty_two_week": [],
        "trending": []
    }

    for symbol in nifty_100_stocks:
        try:
            stock = yf.Ticker(symbol)
            hist = stock.history(period="1y")

            if hist.empty or len(hist) < 2:
                continue

            latest = hist.iloc[-1]
            prev_close = hist.iloc[-2]['Close']
            volume = latest['Volume']
            close_price = latest['Close']
            high_52 = hist['High'].max()
            low_52 = hist['Low'].min()

            avg_volume_7d = hist['Volume'].tail(7).mean()
            volume_spike_ratio = volume / avg_volume_7d if avg_volume_7d else 0
            price_change_pct = ((close_price - prev_close) / prev_close) * 100

            # 1. Most Active Stocks
            data["most_active"].append({
                "symbol": symbol,
                "price": round(close_price, 2),
                "volume": int(volume),
            })

            # 2. 52-Week High/Low Stocks (within 5% of high/low)
            if abs(close_price - high_52) / high_52 < 0.05 or abs(close_price - low_52) / low_52 < 0.05:
                data["fifty_two_week"].append({
                    "symbol": symbol,
                    "price": round(close_price, 2),
                    "high_52": round(high_52, 2),
                    "low_52": round(low_52, 2),
                })

            # 3. Trending Stocks (volume spike > 1.5x or price movement > 3%)
            if volume_spike_ratio > 1.5 or abs(price_change_pct) > 3:
                data["trending"].append({
                    "symbol": symbol,
                    "price": round(close_price, 2),
                    "volume_spike": round(volume_spike_ratio, 2),
                    "price_change_pct": round(price_change_pct, 2),
                })

        except Exception as e:
            logger.warning(f"Error fetching data for {symbol}: {e}")
            continue

    # Sort and limit top 5 results
    data["most_active"] = sorted(data["most_active"], key=lambda x: x["volume"], reverse=True)[:5]
    data["fifty_two_week"] = data["fifty_two_week"][:5]
    data["trending"] = data["trending"][:5]

    # Handle empty cases for frontend
    if not data["most_active"]:
        data["most_active"].append({"message": "No active stocks found."})

    if not data["fifty_two_week"]:
        data["fifty_two_week"].append({"message": "No 52-week high/low stocks today."})

    if not data["trending"]:
        data["trending"].append({"message": "No trending stocks currently."})

    # Add timestamp
    data["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return JsonResponse(data)       


# ✅ Top Gainers & Losers
logger = logging.getLogger(__name__)
@login_required(login_url='/login')
def get_top_gainers_losers(request):
    try:
        # List of NSE stock tickers
        tickers = [
    # Banking & Finance (50)
    "HDFCBANK.NS", "ICICIBANK.NS", "SBIN.NS", "KOTAKBANK.NS", "AXISBANK.NS",
    "INDUSINDBK.NS", "IDFCFIRSTB.NS", "PNB.NS", "BANKBARODA.NS", "FEDERALBNK.NS",
    "CANBK.NS", "RBLBANK.NS", "IDBI.NS", "UCOBANK.NS", "YESBANK.NS",
    "BANDHANBNK.NS", "AUROPHARMA.NS", "MUTHOOTFIN.NS", 
    
    # Information Technology (50)
    "TCS.NS", "INFY.NS", "WIPRO.NS", "HCLTECH.NS", "LTIM.NS",
    "TECHM.NS", "MPHASIS.NS", "COFORGE.NS", "PERSISTENT.NS", "OFSS.NS",
    "MINDTREE.NS", "LTI.NS", "SONATSOFTW.NS", "ZENSARTECH.NS",
    
    # Fast-Moving Consumer Goods (50)
    "HINDUNILVR.NS", "ITC.NS", "NESTLEIND.NS", "BRITANNIA.NS", "DABUR.NS",
    "MARICO.NS", "TATACONSUM.NS", "COLPAL.NS", "GODREJCP.NS", "EMAMILTD.NS",
    "RADICO.NS", "HUL.NS", "JUBLPHARM.NS", "GILLETTE.NS", "VENKEYS.NS",
    "VARUNBEV.NS", "AVANTIFEED.NS", "JUBILANT.NS", "KRBL.NS", 
    # Pharmaceuticals (50)
    "SUNPHARMA.NS", "CIPLA.NS", "DRREDDY.NS", "DIVISLAB.NS", "LUPIN.NS",
    "BIOCON.NS", "ALKEM.NS", "AUROPHARMA.NS", "TORNTPHARM.NS", "ABBOTINDIA.NS",
    "IPCALAB.NS", "GLENMARK.NS", "ALEMBICLTD.NS", "FDC.NS", "NATCO.NS",
    "WOCKPHARMA.NS", "JUBLPHARM.NS", "ERIS.NS", "METROPOLIS.NS", "NEULANDLAB.NS",
]


        # Fetch stock data
        data = yf.download(tickers, period="1d", interval="1h")

        if data.empty:
            return JsonResponse({"error": "No data fetched"}, status=500)

        # Extract Close prices if MultiIndex
        if isinstance(data.columns, pd.MultiIndex):
            data = data.xs('Close', axis=1, level=0)

        # Calculate percentage change
        pct_change = data.pct_change().iloc[-1] * 100  # Latest percentage change
        pct_change = pct_change.dropna().round(2)  # Round to 2 decimal places

        # Get top 5 gainers and losers
        top_gainers = pct_change.nlargest(5).reset_index().values.tolist()
        top_losers = pct_change.nsmallest(5).reset_index().values.tolist()

        # Convert to structured JSON response
        return JsonResponse({
            "top_gainers": [{"symbol": g[0], "change": f"{g[1]:.2f}%"} for g in top_gainers],
            "top_losers": [{"symbol": l[0], "change": f"{l[1]:.2f}%"} for l in top_losers]
        })

    except Exception as e:
        logger.error(f"❌ ERROR in get_top_gainers_losers: {str(e)}")
        return JsonResponse({"error": "Internal Server Error"}, status=500)



# ✅ Market Section sip gainers and losers
@login_required(login_url='/login')
def market_update(request):
    # URL for SIP data
    url = "https://www.moneycontrol.com/mutual-funds/best-funds/"

    # Fetch Data
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    # Extract Table Data (Modify based on actual structure)
    funds_data = []
    rows = soup.find_all("tr")

    for row in rows[1:]:  # Skip header
        cols = row.find_all("td")
        if len(cols) >= 3:
            fund_name = cols[0].text.strip()
            latest_nav = cols[1].text.strip().replace(',', '')  # Remove commas
            change = cols[2].text.strip().replace(',', '').replace('%', '')  # Remove commas & %

            funds_data.append({
                "fund_name": fund_name,
                "nav": latest_nav,
                "change": change
            })

    # Convert 'change' values to float for sorting
    try:
        funds_data.sort(key=lambda x: float(x["change"]), reverse=True)
    except ValueError as e:
        print("Error in conversion:", e)

    # Get top 5 gainers & losers
    top_5_gainers = funds_data[:5]
    top_5_losers = funds_data[-5:]

    return render(request, "market_update.html", {
        "top_5_gainers": top_5_gainers,
        "top_5_losers": top_5_losers
    })


import requests
import random
from django.http import JsonResponse


FRED_API_KEY = "65b4fd45038695649ff36d26a698c216"

# 🟢 **Economic Data Fetch Function**
@login_required(login_url='/login')
def get_economic_events(request):
    try:
        # ✅ **FRED API Endpoint for Economic Data**  
        url = f"https://api.stlouisfed.org/fred/releases?api_key={FRED_API_KEY}&file_type=json"
        
        # ✅ **API Se Data Fetch Karna**  
        response = requests.get(url)
        response_data = response.json()

        # ✅ **Extract Required Data**  
        economic_events = []
        if "releases" in response_data:
            for release in response_data["releases"]:
                economic_events.append({
                    "date": release["realtime_start"],   
                    "event": release["name"],           
                    "value": "N/A",  
                    "end_date": release.get("realtime_end", "N/A"),  
                    "press_release": release.get("press_release", False),  
                    "link": release.get("link", "#")  
                })

        # ✅ **Random 5 Events for Display**
        random_events = random.sample(economic_events, min(8, len(economic_events)))

        return JsonResponse({
            "economic_events": economic_events,  # 🔵 Full Data (Report Ke Liye)
            "random_events": random_events       # 🟢 Random 5 Events (Frontend Ke Liye)
        })
    
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    
    

from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
import requests
from textwrap import wrap

LOGO_PATH = "static/images/auditor.png"
@login_required(login_url='/login')
def download_economic_report(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Draw first page header
    draw_header(p, width, height)
    y = height - 120
    
    # Fetch Data
    try:
        response = requests.get("http://127.0.0.1:8000/dashboard/get-economic-events/")
        data = response.json().get("economic_events", [])
    except Exception as e:
        data = []

    # Table Configurations
    left_margin = 0.5 * inch
    right_margin = 0.5 * inch
    x_start = left_margin
    table_width = width - (left_margin + right_margin)
    col_widths = [1.0 * inch, 2.6 * inch, 1.0 * inch, 0.7 * inch, 3.2 * inch]
    headers = ["Date", "Event", "End Date", "Yes/No", "Link"]

    # Draw Table Header
    y = draw_table_header(p, x_start, y, table_width, col_widths, headers)

    # Print Table Data
    for row in data:
        if y < 70:
            add_footer(p, width)
            p.showPage()
            draw_header(p, width, height)  # Ensure same header on new pages
            y = height - 100
            y = draw_table_header(p, x_start, y, table_width, col_widths, headers)

        event_text = wrap(row.get("event", "N/A"), width=40)
        link_text = wrap(row.get("link", "N/A"), width=33)
        max_lines = max(len(event_text), len(link_text))
        
        p.rect(x_start, y - (max_lines-1) * 10, table_width, 20 + (max_lines-1) * 10, stroke=1, fill=0)
        x_pos = x_start
        
        for i, key in enumerate(["date", "event", "end_date", "press_release", "link"]):
            value = row.get(key, "N/A")
            if key == "press_release":
                value = "Yes" if value else "No"
            text_lines = event_text if key == "event" else link_text if key == "link" else [value]
            
            for j, line in enumerate(text_lines):
                p.drawString(x_pos + 5, y - (j * 10) + 6, line)
            x_pos += col_widths[i]
        
        y -= 20 + (max_lines-1) * 10
    
    add_footer(p, width)
    p.showPage()
    p.save()
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="economic_calendar_report.pdf"'
    return response

# Draw Header

def draw_header(p, width, height):
    try:
        p.drawImage(LOGO_PATH, 40, height - 70, width=80, height=30)
    except:
        pass  
    p.setFont("Helvetica-Bold", 10)  # Unified font size across all pages
    p.drawCentredString(width / 2, height - 60, "Economic Calendar Report")
    p.setFont("Helvetica", 8)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    p.drawCentredString(width / 2, height - 75, "Latest Economic Events & Insights")
    p.setStrokeColorRGB(0, 0, 0)
    p.setLineWidth(1)
    p.line(40, height - 85, width - 40, height - 85)
    p.setFillColorRGB(0, 0, 0)

# Draw Table Header

def draw_table_header(p, x_start, y, table_width, col_widths, headers):
    p.setFont("Helvetica-Bold", 8)
    p.setFillColor(colors.lightgrey)
    p.rect(x_start, y, table_width, 20, fill=1, stroke=1)
    p.setFillColor(colors.black)
    x_pos = x_start
    
    for i, header in enumerate(headers):
        p.drawString(x_pos + 5, y + 6, header)
        p.line(x_pos, y, x_pos, y - 500)
        x_pos += col_widths[i]
    p.line(x_start + table_width, y, x_start + table_width, y - 500)
    
    return y - 20

# Add Footer

def add_footer(p, width):
    p.setFont("Helvetica", 6)
    p.setFillColorRGB(0.5, 0.5, 0.5)
    footer_text = "InvestIQ - Smart Investment Insights | www.investiq.com | support@investiq.com"
    p.drawCentredString(width / 2, 40, footer_text)




# ✅ News Updates
# 🟢 **Finnhub API for News Updates** (Replace with your actual API key)
@login_required(login_url='/login')
def get_news_updates(request):
    try:
        # 🎯 **(Option 1) Finnhub API**
        FINNHUB_API = "https://finnhub.io/api/v1/news?category=general&token=cvhd29pr01qrtb3nse20cvhd29pr01qrtb3nse2g"

        # 🟢 **Choose API**
        API_URL = FINNHUB_API  

        # 🔹 Fetch data from API
        response = requests.get(API_URL)
        data = response.json()

        # 🔹 **Extract News Details Properly**
        latest_news = []
        if isinstance(data, list):  # Finnhub API returns a list
            for article in data[:6]:  # Only get first 10 articles
                latest_news.append({
                    "title": article.get("headline", "No Title"),
                    "description": article.get("summary", "No description available."),
                })

        # 🎯 **Unread news count**
        unread_count = len(latest_news)

        # 🎯 **Return response**
        return JsonResponse({"unread_count": unread_count, "news": latest_news})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



# ✅ Upcoming IPOs
# 🟢 **Finnhub API for Upcoming IPOs** (Replace with your actual API key)
# Set up logging
logger = logging.getLogger(__name__)
@login_required(login_url='/login')
def upcoming_ipos(request):
    api_key = settings.FINNHUB_API_KEY

    # Dynamic date range: today to 30 days ahead
    # from_date = request.GET.get('from', datetime.today().strftime('%Y-%m-%d'))
    # to_date = request.GET.get('to', (datetime.today() + timedelta(days=30)).strftime('%Y-%m-%d'))
    from_date = (datetime.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    to_date = datetime.today().strftime('%Y-%m-%d')
    


    url = f'https://finnhub.io/api/v1/calendar/ipo?from={from_date}&to={to_date}&token={api_key}'

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()

        # Check if IPO data is available
        ipo_data = data.get('ipoCalendar', [])

        if not ipo_data:
            return JsonResponse({'message': 'No upcoming IPOs found.'}, status=204)

        # Optionally: filter or format the IPO data (e.g., company name and date)
        formatted_ipos = [
            {
                'name': item.get('name'),
                'symbol': item.get('symbol'),
                'exchange': item.get('exchange'),
                'date': item.get('date'),
                'numberOfShares': item.get('numberOfShares'),
                'price': item.get('price')
            }
            for item in ipo_data
        ]

        return JsonResponse({'upcoming_ipos': formatted_ipos}, safe=False)

    except requests.exceptions.RequestException as e:
        logger.error(f'Error fetching IPO data: {e}')
        return JsonResponse({'error': 'Failed to fetch IPO data. Please try again later.'}, status=500)
    except ValueError as e:
        logger.error(f'Error parsing JSON: {e}')
        return JsonResponse({'error': 'Invalid data received from API.'}, status=500)





















@login_required(login_url='/')
def report_view(request):
    return render(request, 'report.html', {
        'user_name': request.user.get_full_name() or request.user.username,
        'today': timezone.now().date()
    })


from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from .models import UserProfile, IncomeTracker, Expense
from datetime import timedelta
from django.db.models import Sum
from reportlab.lib.colors import HexColor
from .helpers import wrap_text, get_income_analysis, generate_detailed_income_summary, get_expense_analysis, generate_detailed_expense_summary,  get_combined_financial_analysis, generate_detailed_combined_summary, generate_closing_message
import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

import matplotlib.pyplot as plt
from io import BytesIO
from reportlab.lib.utils import ImageReader



LOGO_PATH = "static/images/auditor.png"

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_pages = []

    def showPage(self):
        self._saved_pages.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        if dict(self.__dict__) not in self._saved_pages:
            self._saved_pages.append(dict(self.__dict__))

        total_pages = len(self._saved_pages)
        for i, page in enumerate(self._saved_pages):
            self.__dict__.update(page)
            draw_footer(self, self._pagesize[0], self._pagesize[1], i + 1, total_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

@login_required(login_url='/login')
def generate_pdf_report(request):
    user = request.user
    name = user.get_full_name() or user.username
    email = user.email
    today = datetime.today().date()

    try:
        profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        profile = None

    mobile = profile.mobile if profile and profile.mobile else "N/A"
    savings = profile.savings if profile and profile.savings else 0
    existing_investments = profile.existing_investments if profile and profile.existing_investments else 0

    duration = request.GET.get("duration", "overall").lower()


    if duration == "monthly":
        start_date = today.replace(day=1)
        report_label = "Monthly"
    elif duration == "annual":
        start_date = today.replace(month=1, day=1)
        report_label = "Annual"
    else:
        start_date = None
        report_label = "Overall"
    
    income_qs = IncomeTracker.objects.filter(user=user)
    expense_qs = Expense.objects.filter(user=user)
    if start_date:
        income_qs = income_qs.filter(date__gte=start_date, date__lte=today)
        expense_qs = expense_qs.filter(date__gte=start_date, date__lte=today)

    income = income_qs.aggregate(total=Sum('amount'))['total'] or 0
    expense = expense_qs.aggregate(total=Sum('amount'))['total'] or 0

    buffer = BytesIO()
    p = NumberedCanvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 130

    draw_header(p, width, height, today, name, email, mobile, report_label)

    # Section: Financial Overview Table
    p.setFont("Helvetica-Bold", 16)
    p.setFillColor(HexColor("#333333"))
    p.drawCentredString(width / 2, y, "User Financial Details")
    y -= 30

    columns = ["Total Income", "Total Savings", "Existing Investments", "Total Expenses"]
    values = [income, savings, existing_investments, expense]
    table_left, table_right = 50, width - 50
    col_width = (table_right - table_left) / len(columns)

    header_y = y
    value_y = y - 25

    p.setFont("Helvetica-Bold", 12)
    p.setFillColor(HexColor("#555555"))
    for i, col in enumerate(columns):
        x = table_left + i * col_width + col_width / 2
        p.drawCentredString(x, header_y, col)

    p.setFont("Helvetica", 12)
    p.setFillColor(HexColor("#000000"))
    for i, val in enumerate(values):
        x = table_left + i * col_width + col_width / 2
        p.drawCentredString(x, value_y, f"Rs.{val:,.2f}")

    p.setStrokeColor(HexColor("#cccccc"))
    p.setLineWidth(1)
    p.line(table_left, header_y + 15, table_right, header_y + 15)
    p.line(table_left, header_y - 10, table_right, header_y - 10)
    p.line(table_left, value_y - 10, table_right, value_y - 10)
    for i in range(len(columns) + 1):
        x = table_left + i * col_width
        p.line(x, header_y + 15, x, value_y - 10)

    p.rect(table_left, value_y - 10, table_right - table_left, (header_y + 15) - (value_y - 10), stroke=1, fill=0)
    y = value_y - 50







    # Section: Income Analysis Summary

    income_analysis = get_income_analysis(user, period=duration)
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width / 2, y, "Income Analysis Summary")
    y -= 15

    
    # Labels and values
    labels = [
        "Total Income",
        "Avg Monthly Income",
        "Passive Income Ratio",
        "Avg Entry Frequency",
        "Highest Source",
        "Lowest Source"
    ]
    values = [
        f"Rs.{income_analysis['total_income']:,.2f}",
        f"Rs.{income_analysis['average_monthly_income']:,.2f}",
        f"{income_analysis['highest_source']['source']} (Rs.{income_analysis['highest_source']['total']:,.2f})" if income_analysis['highest_source'] else "N/A",
        f"{income_analysis['lowest_source']['source']} (Rs.{income_analysis['lowest_source']['total']:,.2f})" if income_analysis['lowest_source'] else "N/A",
        f"{income_analysis['passive_ratio']:.2f}%",
        f"{income_analysis['avg_entry_frequency']:.1f}"
        
    ]
    
    # Grid layout (2 rows × 3 columns)
    cols = 3
    rows = 2
    col_width = (width - 120) / cols
    row_height = 40
    
    for row in range(rows):
        for col in range(cols):
            idx = row * cols + col
            x = 60 + col * col_width
            box_y = y - row * row_height
    
            # Draw outer box
            p.setStrokeColor(HexColor("#cccccc"))
            p.rect(x, box_y - row_height, col_width, row_height, stroke=1, fill=0)
    
            # Label
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(HexColor("#555555"))
            p.drawString(x + 8, box_y - 12, labels[idx])
    
            # Value
            p.setFont("Helvetica", 11)
            p.setFillColor(HexColor("#000000"))
            p.drawString(x + 8, box_y - 27, values[idx])
            
     # Update y after grid
    y -= row_height * rows + 20

    
    # Income Source Pie Chart
    plt.figure(figsize=(4.5, 4))
    plt.pie(
        income_analysis['source_amounts'],
        labels=income_analysis['source_labels'],
        autopct='%1.1f%%',
        startangle=140,
        colors=plt.cm.Paired.colors
    )
    plt.title(
        "Income Source Breakdown",
        fontsize=14,       # size of the font
        fontweight='bold', # bold text
        color='#333333',   # dark gray color
        pad=15             # padding above the plot
    )
    plt.tight_layout()
    
    # Save and draw pie chart
    pie_buffer = BytesIO()
    plt.savefig(pie_buffer, format='PNG')
    plt.close()
    pie_buffer.seek(0)
    pie_img = ImageReader(pie_buffer)
    p.drawImage(pie_img, 60, y - 220, width=(width - 120) / 2, height=200)

    # Monthly Entry Frequency Bar Chart
    plt.figure(figsize=(4.5, 4))
    plt.bar(income_analysis['bar_labels'], income_analysis['bar_values'], color="#4CAF50")
    plt.xticks(rotation=45, ha='right')
    plt.title(
        "Monthly Income Entry Frequency",
        fontsize=14,       # size of the font
        fontweight='bold', # bold text
        color='#333333',   # dark gray color
        pad=15             # padding above the plot
    )

    plt.xlabel("Month")
    plt.ylabel("Entries")
    plt.tight_layout()
    
    # Save and draw bar chart
    bar_buffer = BytesIO()
    plt.savefig(bar_buffer, format='PNG')
    plt.close()
    bar_buffer.seek(0)
    bar_img = ImageReader(bar_buffer)
    p.drawImage(bar_img, 60 + (width - 120) / 2, y - 220, width=(width - 120) / 2, height=200)


    y -= 220 + 30  # Adjust y after both graphs

        # ✅ Fix y position after all graphs
    # y = current_y - 30
    
    
    # Now start drawing INCOME text summary...
    analysis_data = get_income_analysis(user, period=duration)
    detailed_text = generate_detailed_income_summary(analysis_data)
    
    lines = detailed_text.split('\n')
    normal_font = ("Helvetica", 10)
    bold_font = ("Helvetica-Bold", 12)
    heading_font = ("Helvetica-Bold", 12)
    
    line_height_normal = 12
    line_height_heading = 16
    
    space_after_paragraph = 1
    space_before_heading = 10
    bottom_margin = 70
    max_width = width - 120  # 60 padding dono side
    
    p.setFillColor(HexColor("#000000"))
    
    for paragraph_line in lines:
        stripped_line = paragraph_line.strip()
        if not stripped_line:
            y -= space_after_paragraph // 2
            continue
    
        # Heading detection: line fully uppercase or contains key words like "summary", "analysis", etc.
        if stripped_line.isupper():
            y -= space_before_heading
            font_name, font_size = heading_font if len(stripped_line) < 100 else bold_font
            line_height = line_height_heading
        else:
            font_name, font_size = normal_font
            line_height = line_height_normal
    
        p.setFont(font_name, font_size)
    
        wrapped_lines = wrap_text(stripped_line, max_width, font_name, font_size)
        for line in wrapped_lines:
            if y < bottom_margin:
                p.showPage()
                draw_header(p, width, height, today, name, email, mobile, report_label)
                y = height - 130
                p.setFillColor(HexColor("#000000"))
                p.setFont(font_name, font_size)
            p.drawString(60, y, line)
            y -= line_height
    
        y -= space_after_paragraph
    

    p.showPage()










    # Section: Expense Analysis Summary
    draw_header(p, width, height, today, name, email, mobile, report_label)
    y = height - 130

    expense_analysis = get_expense_analysis(user, period=duration)
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width / 2, y, "Expense Analysis Summary")
    y -= 15
    
    labels = [
        "Total Expense",
        "Highest Spending Category",
        "Lowest Spending Category",
        "Avg Entry Frequency",
        "Budget Used %",
        "Monthly Growth %"
    ]
    
    # For Budget Used %, hum average % use kar lenge sab categories ke budget usage ka
    if expense_analysis['budget_comparison']:
        avg_budget_used = sum(item['percentage_used'] for item in expense_analysis['budget_comparison']) / len(expense_analysis['budget_comparison'])
    else:
        avg_budget_used = 0
    
    # Monthly growth agar data hai to last month ka growth lenge, warna 0
    monthly_growth = expense_analysis['monthly_growth'][-1] if expense_analysis['monthly_growth'] else 0
    
    values = [
        f"Rs.{expense_analysis['total_expense']:,.2f}",
        f"{expense_analysis['highest_category']['category']} (Rs.{expense_analysis['highest_category']['total']:,.2f})",
        f"{expense_analysis['lowest_category']['category']} (Rs.{expense_analysis['lowest_category']['total']:,.2f})",
        f"{expense_analysis['avg_entry_frequency']:.1f}",
        f"{avg_budget_used:.2f}%",
        f"{monthly_growth:.2f}%"
    ]
    
    cols = 3
    rows = 2
    col_width = (width - 120) / cols
    row_height = 40
    
    for row in range(rows):
        for col in range(cols):
            idx = row * cols + col
            x = 60 + col * col_width
            box_y = y - row * row_height
    
            # Draw outer box
            p.setStrokeColor(HexColor("#cccccc"))
            p.rect(x, box_y - row_height, col_width, row_height, stroke=1, fill=0)
    
            # Label
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(HexColor("#555555"))
            p.drawString(x + 8, box_y - 12, labels[idx])
    
            # Value
            p.setFont("Helvetica", 11)
            p.setFillColor(HexColor("#000000"))
            p.drawString(x + 8, box_y - 27, values[idx])
    
    # Update y after grid
    y -= row_height * rows + 20

    
    # Graph dimensions and positioning
    graph_width = width - 100
    graph_height = 240
    x_pos = 50
    current_y = y  # Start from current y
    

    # --------- 1. Monthly Expense Trend ---------
    plt.figure(figsize=(7, 3.5))
    plt.plot(expense_analysis['bar_labels'], expense_analysis['bar_values'], marker='o', color='#1f77b4')
    plt.xticks(rotation=45, ha='right')
    plt.title("Monthly Expense Trend", fontsize=14, fontweight='bold', color='#333333', pad=15)
    plt.xlabel("Month")
    plt.ylabel("Amount (Rs.)")
    plt.tight_layout()
    buf1 = BytesIO()
    plt.savefig(buf1, format='PNG')
    plt.close()
    buf1.seek(0)
    
    # --------- 2. Spending by Weekday (replace Monthly Growth %) ---------
    plt.figure(figsize=(7, 3.5))
    plt.bar(expense_analysis['weekday_labels'], expense_analysis['weekday_data'], color='#ffa500')
    plt.title("Spending by Weekday", fontsize=14, fontweight='bold', color='#333333', pad=15)
    plt.xlabel("Day of Week")
    plt.ylabel("Amount (Rs.)")
    plt.tight_layout()
    buf2 = BytesIO()
    plt.savefig(buf2, format='PNG')
    plt.close()
    buf2.seek(0)
    
    # --------- 3. Category-wise Expense Breakdown ---------
    plt.figure(figsize=(7, 3.5))
    plt.pie(
        expense_analysis['category_totals'],
        labels=expense_analysis['category_labels'],
        autopct='%1.1f%%',
        startangle=140,
        colors=plt.cm.Paired.colors
    )
    plt.title("Category-wise Expense Breakdown", fontsize=14, fontweight='bold', color='#333333', pad=15)
    plt.tight_layout()
    buf3 = BytesIO()
    plt.savefig(buf3, format='PNG')
    plt.close()
    buf3.seek(0)
    
    # --------- 4. Limit vs Actual ---------
    budget_data = expense_analysis['budget_comparison']
    if budget_data:
        labels = [item['category'] for item in budget_data]
        actual = [item['actual'] for item in budget_data]
        limit = [item['limit'] for item in budget_data]
    
        x = range(len(labels))
        plt.figure(figsize=(7, 3.5))
        plt.bar(x, actual, width=0.4, label='Actual', align='center', color='#2ca02c')
        plt.bar([i + 0.4 for i in x], limit, width=0.4, label='Limit', align='center', color='#d62728')
        plt.xticks([i + 0.2 for i in x], labels, rotation=45, ha='right')
        plt.title("Limit vs Actual Spending", fontsize=14, fontweight='bold', color='#333333', pad=15)
        plt.xlabel("Category")
        plt.ylabel("Amount (Rs.)")
        plt.legend()
        plt.tight_layout()
        buf4 = BytesIO()
        plt.savefig(buf4, format='PNG')
        plt.close()
        buf4.seek(0)
    else:
        buf4 = BytesIO()  # Empty image buffer if no budget data
    
    # --------- Draw graphs vertically ---------
    graphs = [buf1, buf2, buf3, buf4]  # Note: buf2 is now Spending by Weekday, buf3 is Category-wise Breakdown
    titles = [
        "Monthly Expense Trend",
        "Spending by Weekday",
        "Category-wise Breakdown",
        "Limit vs Actual"
    ]
    
    for i, buf in enumerate(graphs):
        if i == 2:
            p.showPage()
            draw_header(p, width, height, today, name, email, mobile, report_label)
            current_y = height - 130
    
        current_y -= graph_height
        p.drawImage(ImageReader(buf), x_pos, current_y, width=graph_width, height=graph_height)
        current_y -= 20
    
    # Update y position after drawing graphs
    y = current_y - 30


    
    # Now start drawing text summary...
    # Now start drawing EXPENSE text summary...
    analysis_data = get_expense_analysis(user, period=duration)
    detailed_text = generate_detailed_expense_summary(analysis_data)
    
    lines = detailed_text.split('\n')
    normal_font = ("Helvetica", 10)
    bold_font = ("Helvetica-Bold", 12)
    heading_font = ("Helvetica-Bold", 12)
    
    line_height_normal = 12
    line_height_heading = 16
    
    space_after_paragraph = 1
    space_before_heading = 10
    
    p.setFillColor(HexColor("#000000"))
    
    for paragraph_line in lines:
        stripped_line = paragraph_line.strip()
        if not stripped_line:
            y -= space_after_paragraph // 2
            continue
    
        if stripped_line.isupper():
            y -= space_before_heading
            font_name, font_size = heading_font if len(stripped_line) < 100 else bold_font
            line_height = line_height_heading
        else:
            font_name, font_size = normal_font
            line_height = line_height_normal
    
        p.setFont(font_name, font_size)
    
        wrapped_lines = wrap_text(stripped_line, max_width, font_name, font_size)
        for line in wrapped_lines:
            if y < bottom_margin:
                p.showPage()
                draw_header(p, width, height, today, name, email, mobile, report_label)
                y = height - 130
                p.setFillColor(HexColor("#000000"))
                p.setFont(font_name, font_size)
            p.drawString(60, y, line)
            y -= line_height
    
        y -= space_after_paragraph
    

    p.showPage()





    # Section: Combined Analysis Summary

    draw_header(p, width, height, today, name, email, mobile, report_label)
    y = height - 130
    
    combined_analysis = get_combined_financial_analysis(user, period=duration)
   
    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(HexColor("#333333"))
    p.drawCentredString(width / 2, y, "Combined Analysis Summary")
    y -= 20  # Slightly more gap for heading
    
    labels = [
        "Total Income",
        "Total Expense",
        "Inc / Exp Ratio",
        "Budget Accuracy Score"
    ]
    
    total_income = combined_analysis.get('total_income', 0)
    total_expense = combined_analysis.get('total_expense', 0)
    income_expense_ratio = combined_analysis.get('income_expense_ratio', '∞')
    budget_accuracy = combined_analysis.get('budget_accuracy_score', 0)
    
    values = [
        f"Rs. {total_income:,.2f}",
        f"Rs. {total_expense:,.2f}",
        f"{income_expense_ratio}",
        f"{budget_accuracy if budget_accuracy is not None else 'N/A'}%"
    ]
    
    cols = 4
    col_width = (width - 120) / cols
    row_height = 55  # Increased height
    box_y = y
    
    p.setStrokeColor(HexColor("#cccccc"))
    p.setLineWidth(1)
    
    # Draw boxes, labels, and values
    for col in range(cols):
        x = 60 + col * col_width
        
        # Box border
        p.rect(x, box_y - row_height, col_width, row_height, stroke=1, fill=0)
        
        # Label
        p.setFont("Helvetica-Bold", 11)
        p.setFillColor(HexColor("#555555"))
        p.drawCentredString(x + col_width / 2, box_y - 15, labels[col])
        
        # Value
        p.setFont("Helvetica", 13)
        p.setFillColor(HexColor("#000000"))
        p.drawCentredString(x + col_width / 2, box_y - 38, values[col])
    
    # Update y for next section
    y -= row_height + 20
    



    # -------- Setup basic variables --------
    graph_width = width - 100
    graph_height = 240
    x_pos = 50
    current_y = y
    
    graphs = []  # Will store chart buffers in order
    
    # --------- 1. Monthly Income vs Expense ---------
    monthly_data = combined_analysis.get('monthly_cash_flow', {})
    if monthly_data:
        months = list(monthly_data.keys())
        monthly_income = [val['income'] for val in monthly_data.values()]
        monthly_expense = [val['expense'] for val in monthly_data.values()]
    
        plt.figure(figsize=(7, 3.5))
        plt.plot(months, monthly_income, label="Income", color='green', marker='o')
        plt.plot(months, monthly_expense, label="Expense", color='red', marker='o')
        plt.xticks(rotation=45)
        plt.title("Monthly Income vs Expense", fontsize=14, fontweight='bold')
        plt.xlabel("Month")
        plt.ylabel("Amount")
        plt.legend()
        plt.tight_layout()
        buf1 = BytesIO()
        plt.savefig(buf1, format='PNG')
        plt.close()
        buf1.seek(0)
        graphs.append(buf1)
    
    # --------- 2. Budget vs Actual Spending ---------
    budget_data = combined_analysis.get("budget_vs_actual", [])
    if budget_data:
        labels = [item['category'] for item in budget_data]
        actual = [item['actual'] for item in budget_data]
        budget = [item['budget'] for item in budget_data]
    
        x = range(len(labels))
        plt.figure(figsize=(7, 3.5))
        plt.bar(x, actual, width=0.4, label="Actual", color='skyblue', align='center')
        plt.bar([i + 0.4 for i in x], budget, width=0.4, label="Budget", color='salmon', align='center')
        plt.xticks([i + 0.2 for i in x], labels, rotation=45)
        plt.title("Budget vs Actual Spending", fontsize=14, fontweight='bold')
        plt.xlabel("Category")
        plt.ylabel("Amount")
        plt.legend()
        plt.tight_layout()
        buf2 = BytesIO()
        plt.savefig(buf2, format='PNG')
        plt.close()
        buf2.seek(0)
        graphs.append(buf3)
    
    # --------- 3. Ideal vs Actual Spend-Save-Invest Ratio ---------
    ideal_actual = combined_analysis.get('ideal_vs_actual_ratio', {})
    ideal = ideal_actual.get('ideal', {})
    actual = ideal_actual.get('actual', {})
    
    buf3a = buf3b = None
    if ideal and actual:
        # Ideal pie chart
        plt.figure(figsize=(3.5, 3.5))
        plt.pie(ideal.values(), labels=ideal.keys(), autopct='%1.1f%%', startangle=140, colors=plt.cm.Set3.colors)
        plt.title("Ideal Spend-Save-Invest", fontsize=12, fontweight='bold')
        buf3a = BytesIO()
        plt.savefig(buf3a, format='PNG')
        plt.close()
        buf3a.seek(0)
    
        # Actual pie chart
        plt.figure(figsize=(3.5, 3.5))
        plt.pie(actual.values(), labels=actual.keys(), autopct='%1.1f%%', startangle=140, colors=plt.cm.Pastel2.colors)
        plt.title("Actual Spend-Save-Invest", fontsize=12, fontweight='bold')
        buf3b = BytesIO()
        plt.savefig(buf3b, format='PNG')
        plt.close()
        buf3b.seek(0)
    
    # --------- Draw all graphs into PDF ---------
    for i, buf in enumerate(graphs):
        if i == 2:  # New page after 2nd graph now
            p.showPage()
            draw_header(p, width, height, today, name, email, mobile, report_label)
            current_y = height - 130
    
        current_y -= graph_height
        p.drawImage(ImageReader(buf), x_pos, current_y, width=graph_width, height=graph_height)
        current_y -= 18
    
    # --------- Draw Pie Charts if Available on NEW PAGE ---------
    if buf3a and buf3b:
        p.showPage()  # Start new page for pie charts
        draw_header(p, width, height, today, name, email, mobile, report_label)
        current_y = height - 130  # reset y position for new page
    
        current_y -= graph_height
        p.drawImage(ImageReader(buf3a), x_pos, current_y, width=(graph_width // 2) - 10, height=graph_height)
        p.drawImage(ImageReader(buf3b), x_pos + (graph_width // 2) + 10, current_y, width=(graph_width // 2) - 10, height=graph_height)
        y = current_y - 30

    



    combined_data = get_combined_financial_analysis(user, period=duration)
    combined_text = generate_detailed_combined_summary(combined_data)    

    lines = combined_text.split('\n')
    normal_font = ("Helvetica", 10)
    bold_font = ("Helvetica-Bold", 12)
    heading_font = ("Helvetica-Bold", 12)    

    line_height_normal = 12   # thoda kam kiya from 14
    line_height_heading = 16  # thoda kam kiya from 18    

    space_after_paragraph = 1    # kam kiya from 10
    space_before_heading = 10    # 2 times space_after_paragraph = 2*6=12    

    p.setFillColor(HexColor("#000000"))    

    for paragraph_line in lines:
        stripped_line = paragraph_line.strip()
        if not stripped_line:
            y -= space_after_paragraph // 2  # half space for empty lines (optional)
            continue    

        # Check if heading
        if stripped_line.isupper():
            # Add extra space BEFORE heading (2 parts)
            y -= space_before_heading    

            font_name, font_size = heading_font if len(stripped_line) < 100 else bold_font
            line_height = line_height_heading
        else:
            font_name, font_size = normal_font
            line_height = line_height_normal    

        p.setFont(font_name, font_size)    

        wrapped_lines = wrap_text(stripped_line, max_width, font_name, font_size)
        for line in wrapped_lines:
            if y < bottom_margin:
                p.showPage()
                draw_header(p, width, height, today, name, email, mobile, report_label)
                y = height - 130
                p.setFillColor(HexColor("#000000"))
                p.setFont(font_name, font_size)
            p.drawString(60, y, line)
            y -= line_height    

        # Add smaller space AFTER paragraph/heading (1 part)
        y -= space_after_paragraph



    p.showPage()
    draw_header(p, width, height, today, name, email, mobile, report_label)

    # Call the function to get the message
    closing_text = generate_closing_message()
    
    # Set font and margins
    p.setFont("Helvetica", 11)
    p.setFillColor(HexColor("#000000"))
    
    from textwrap import wrap
    y = height - 130  # top margin
    x = 60            # left margin
    max_width = width - 120
    
    # Draw the message with wrapped lines
    for para in closing_text.split('\n'):
        wrapped_lines = wrap(para, width=90)  # or use custom wrap_text() if you have
        for line in wrapped_lines:
            if y < 60:  # bottom margin
                p.showPage()
                p.setFont("Helvetica", 11)
                y = height - 100
            p.drawString(x, y, line)
            y -= 15  # line spacing
        y -= 10  # paragraph spacing
    


    p.save()
    buffer.seek(0)
    # p.showPage()
    # draw_header(p, width, height, today, name, email, mobile, report_label)  


    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="finance_report_{report_label.lower()}.pdf"'
    return response 

    
    # p.save()
    # buffer.seek(0)
    
    # response = HttpResponse(buffer, content_type='application/pdf')
    # response['Content-Disposition'] = 'attachment; filename="finance_report.pdf"'
    # return response



    

    # p.save()
    # buffer.seek(0)

    # response = HttpResponse(buffer, content_type='application/pdf')
    # response['Content-Disposition'] = 'attachment; filename="finance_report.pdf"'
    # return response

def draw_header(p, width, height, date, name, email, mobile, report_type):
    try:
        p.drawImage(LOGO_PATH, 40, height - 95, width=70, height=70, mask='auto')
    except Exception:
        pass

    p.setFont("Helvetica-Bold", 18)
    p.drawCentredString(width / 2, height - 40, "PERSONAL FINANCE MANAGEMENT")

    p.setFont("Helvetica-Bold", 14)
    p.setFillColorRGB(0.2, 0.2, 0.2)
    p.drawCentredString(width / 2, height - 55, f"{report_type.upper()} REPORT")

    p.setFont("Helvetica-Bold", 9)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    p.drawCentredString(width / 2, height - 70, "DETAILED INCOME AND EXPENSE REPORT")

    p.setFont("Helvetica", 8)
    p.setFillColorRGB(0, 0, 0)
    p.drawRightString(width - 40, height - 35, f"DATE: {date.strftime('%d %b %Y')}")

    p.setFont("Helvetica-Bold", 8)
    p.setFillColorRGB(0.2, 0.2, 0.2)
    user_info_text = f"Name: {name}    |    Email: {email}    |    Mobile: {mobile}"
    p.drawCentredString(width / 2, height - 92, user_info_text)

    p.setStrokeColorRGB(0, 0, 0)
    p.setLineWidth(1)
    p.line(40, height - 100, width - 40, height - 100)

def draw_footer(p, width, height, page_number, total_pages):
    p.setFont("Helvetica-Bold", 8)
    p.setFillColorRGB(0.3, 0.3, 0.3)
    p.drawRightString(width - 40, 65, f"Page {page_number} of {total_pages}")

    p.setStrokeColorRGB(0, 0, 0)
    p.setLineWidth(1)
    p.line(40, 60, width - 40, 60)

    p.setFont("Helvetica", 7)
    p.setFillColorRGB(0.5, 0.5, 0.5)
    footer_text = (
        "Personal Finance Management - Smart Investment Insights | "
        "www.personalfinancemanager.com | support@personalfinancemanagement.com"
    )
    p.drawCentredString(width / 2, 45, footer_text)





from django.http import HttpResponse
import csv
from datetime import datetime
from .helpers import (
    get_income_analysis,
    get_expense_analysis,
    get_combined_financial_analysis,
)


# === Helpers ===

def format_currency(value):
    try:
        return f"{float(value):.2f}"
    except (ValueError, TypeError):
        return "0.00"

def add_blank_row(writer, count=1):
    for _ in range(count):
        writer.writerow([])

def write_income_section(writer, summary):
    writer.writerow(['Income Summary'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Total Income', format_currency(summary.get('total_income'))])
    writer.writerow(['Average Monthly Income', format_currency(summary.get('average_monthly_income'))])
    writer.writerow(['Passive Income Ratio (%)', summary.get('passive_ratio', 'N/A')])
    writer.writerow(['Average Entry Frequency', summary.get('avg_entry_frequency', 'N/A')])

    highest = summary.get('highest_source') or {'source': 'N/A', 'total': 0}
    lowest = summary.get('lowest_source') or {'source': 'N/A', 'total': 0}
    writer.writerow(['Highest Income Source', f"{highest['source']} - {format_currency(highest['total'])}"])
    writer.writerow(['Lowest Income Source', f"{lowest['source']} - {format_currency(lowest['total'])}"])

    add_blank_row(writer)
    writer.writerow(['Income by Source'])
    writer.writerow(['Source', 'Amount'])
    for source, amount in zip(summary.get('source_labels', []), summary.get('source_amounts', [])):
        writer.writerow([source, format_currency(amount)])

    add_blank_row(writer)
    writer.writerow(['Monthly Income Entry Frequency'])
    writer.writerow(['Month', 'Entries'])
    for month, count in zip(summary.get('bar_labels', []), summary.get('bar_values', [])):
        writer.writerow([month, count])


def write_expense_section(writer, summary, period):
    writer.writerow(['Expense Summary'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Total Expense', format_currency(summary.get('total_expense'))])
    writer.writerow(['Average Monthly Entries', summary.get('avg_entry_frequency', 'N/A')])
    writer.writerow(['Highest Spending Category', f"{summary['highest_category']['category']} - {format_currency(summary['highest_category']['total'])}"])
    writer.writerow(['Lowest Spending Category', f"{summary['lowest_category']['category']} - {format_currency(summary['lowest_category']['total'])}"])

    add_blank_row(writer)
    writer.writerow(['Category-wise Expense Breakdown'])
    writer.writerow(['Category', 'Total Amount', 'Percentage'])
    for label, total, percent in zip(summary.get('category_labels', []), summary.get('category_totals', []), summary.get('category_percentages', [])):
        writer.writerow([label, format_currency(total), f"{round(percent, 2)}%"])

    add_blank_row(writer)
    writer.writerow(['Monthly Expense Trend'])
    writer.writerow(['Month', 'Expense'])
    for month, amount in zip(summary.get('bar_labels', []), summary.get('bar_values', [])):
        writer.writerow([month, format_currency(amount)])

    if summary.get('monthly_growth'):
        add_blank_row(writer)
        writer.writerow(['Monthly Growth (%)'])
        writer.writerow(['Month Index', 'Growth %'])
        for i, growth in enumerate(summary['monthly_growth']):
            writer.writerow([i + 1, f"{growth}%"])

    add_blank_row(writer)
    writer.writerow(['Spending by Weekday'])
    writer.writerow(['Weekday', 'Amount'])
    for day, value in zip(summary.get('weekday_labels', []), summary.get('weekday_data', [])):
        writer.writerow([day, format_currency(value)])

    if period == 'monthly':
        add_blank_row(writer)
        writer.writerow(['Spending by Day of Month'])
        writer.writerow(['Day', 'Amount'])
        for day, amount in zip(summary.get('dom_labels', []), summary.get('dom_data', [])):
            writer.writerow([day, format_currency(amount)])

    add_blank_row(writer)
    writer.writerow(['Budget vs Actual Spending (Monthly Limits)'])
    writer.writerow(['Category', 'Limit', 'Actual', 'Used (%)'])
    for item in summary.get('budget_comparison', []):
        writer.writerow([
            item['category'],
            format_currency(item['limit']),
            format_currency(item['actual']),
            f"{item['percentage_used']}%"
        ])


def write_combined_section(writer, summary):
    writer.writerow(['Combined Financial Overview'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Total Income', format_currency(summary.get('total_income'))])
    writer.writerow(['Total Expense', format_currency(summary.get('total_expense'))])
    writer.writerow(['Income to Expense Ratio', summary.get('income_expense_ratio', 'N/A')])
    writer.writerow(['Budget Accuracy Score', summary.get('budget_accuracy_score', 'N/A')])

    add_blank_row(writer)
    writer.writerow(['Ideal vs Actual Spend-Save-Invest Ratio'])
    writer.writerow(['Category', 'Ideal (%)', 'Actual (%)'])
    ideal_actual = summary.get('ideal_vs_actual_ratio', {'ideal': {}, 'actual': {}})
    for key in ['spend', 'save', 'invest']:
        writer.writerow([
            key.title(),
            ideal_actual['ideal'].get(key, 0),
            ideal_actual['actual'].get(key, 0)
        ])

    add_blank_row(writer)
    writer.writerow(['Budget vs Actual by Category'])
    writer.writerow(['Category', 'Budget', 'Actual', 'Remaining'])
    for item in summary.get('budget_vs_actual', []):
        writer.writerow([
            item['category'],
            format_currency(item['budget']),
            format_currency(item['actual']),
            format_currency(item['remaining'])
        ])

    if 'monthly_cash_flow' in summary:
        add_blank_row(writer)
        writer.writerow(['Monthly Cash Flow'])
        writer.writerow(['Month', 'Income', 'Expense'])
        for month, data in summary['monthly_cash_flow'].items():
            writer.writerow([
                month,
                format_currency(data['income']),
                format_currency(data['expense'])
            ])

    if 'annual_cash_flow' in summary:
        add_blank_row(writer)
        writer.writerow(['Annual Cash Flow'])
        writer.writerow(['Year', 'Income', 'Expense'])
        for year, data in summary['annual_cash_flow'].items():
            writer.writerow([
                year,
                format_currency(data['income']),
                format_currency(data['expense'])
            ])


# === Main CSV View ===

def generate_csv_report(request):
    user = request.user
    period = request.GET.get('duration', 'annual')
    delimiter = '\t' if request.GET.get('format') == 'tsv' else ','
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Fetch analysis
    income_summary = get_income_analysis(user, period)
    expense_summary = get_expense_analysis(user, period)
    combined_summary = get_combined_financial_analysis(user, period)

    # Prepare CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=financial_summary_{period}_{timestamp}.csv'
    writer = csv.writer(response, delimiter=delimiter)

    # Write each section
    write_income_section(writer, income_summary)
    add_blank_row(writer, 2)

    write_expense_section(writer, expense_summary, period)
    add_blank_row(writer, 2)

    write_combined_section(writer, combined_summary)

    return response




from django.http import HttpResponse
import openpyxl
from .helpers import get_income_analysis, get_expense_analysis, get_combined_financial_analysis

def generate_excel_report(request):
    user = request.user
    period = request.GET.get('duration', 'annual')

    income_summary = get_income_analysis(user, period)
    expense_summary = get_expense_analysis(user, period)
    combined_summary = get_combined_financial_analysis(user, period)

    wb = openpyxl.Workbook()

    ### INCOME SUMMARY ###
    ws1 = wb.active
    ws1.title = "Income Summary"
    ws1.append(['Metric', 'Value'])
    ws1.append(['Total Income', income_summary['total_income']])
    ws1.append(['Average Monthly Income', income_summary['average_monthly_income']])
    ws1.append(['Passive Income Ratio (%)', income_summary['passive_ratio']])
    ws1.append(['Average Entry Frequency', income_summary['avg_entry_frequency']])
    highest = income_summary['highest_source']
    lowest = income_summary['lowest_source']
    ws1.append(['Highest Income Source', f"{highest['source']} - {highest['total']}" if highest else 'N/A'])
    ws1.append(['Lowest Income Source', f"{lowest['source']} - {lowest['total']}" if lowest else 'N/A'])
    ws1.append([])
    ws1.append(['Income by Source'])
    ws1.append(['Source', 'Amount'])
    for source, amount in zip(income_summary['source_labels'], income_summary['source_amounts']):
        ws1.append([source, amount])
    ws1.append([])
    ws1.append(['Monthly Entry Frequency'])
    ws1.append(['Month', 'Entries'])
    for month, count in zip(income_summary['bar_labels'], income_summary['bar_values']):
        ws1.append([month, count])

    ### EXPENSE SUMMARY ###
    ws2 = wb.create_sheet(title="Expense Summary")
    ws2.append(['Metric', 'Value'])
    ws2.append(['Total Expense', expense_summary['total_expense']])
    ws2.append(['Average Entry Frequency', expense_summary['avg_entry_frequency']])
    highest = expense_summary['highest_category']
    lowest = expense_summary['lowest_category']
    ws2.append(['Highest Expense Category', f"{highest['category']} - {highest['total']}"])
    ws2.append(['Lowest Expense Category', f"{lowest['category']} - {lowest['total']}"])
    ws2.append([])
    ws2.append(['Category-wise Spending'])
    ws2.append(['Category', 'Amount', 'Percentage (%)'])
    for label, total, percent in zip(expense_summary['category_labels'], expense_summary['category_totals'], expense_summary['category_percentages']):
        ws2.append([label, total, round(percent, 2)])
    ws2.append([])
    ws2.append(['Budget vs Actual (Monthly)'])
    ws2.append(['Category', 'Budget Limit', 'Actual Spend', 'Percentage Used'])
    for item in expense_summary['budget_comparison']:
        ws2.append([item['category'], item['limit'], item['actual'], item['percentage_used']])
    ws2.append([])
    ws2.append(['Monthly Expense Trend'])
    ws2.append(['Month', 'Amount'])
    for label, value in zip(expense_summary['bar_labels'], expense_summary['bar_values']):
        ws2.append([label, value])
    if expense_summary['monthly_growth']:
        ws2.append([])
        ws2.append(['Monthly Growth (%)'])
        ws2.append(['From Month', 'Growth %'])
        for i in range(1, len(expense_summary['bar_labels'])):
            ws2.append([f"{expense_summary['bar_labels'][i-1]} → {expense_summary['bar_labels'][i]}", expense_summary['monthly_growth'][i-1]])
    ws2.append([])
    ws2.append(['Spending by Weekday'])
    ws2.append(['Weekday', 'Total Amount'])
    for day, amt in zip(expense_summary['weekday_labels'], expense_summary['weekday_data']):
        ws2.append([day, amt])
    if expense_summary['dom_data']:
        ws2.append([])
        ws2.append(['Spending by Day of Month'])
        ws2.append(['Day', 'Amount'])
        for day, amt in zip(expense_summary['dom_labels'], expense_summary['dom_data']):
            ws2.append([day, amt])

    ### COMBINED SUMMARY ###
    ws3 = wb.create_sheet(title="Combined Summary")
    ws3.append(['Metric', 'Value'])
    ws3.append(['Total Income', combined_summary['total_income']])
    ws3.append(['Total Expense', combined_summary['total_expense']])
    ws3.append(['Income-Expense Ratio', combined_summary['income_expense_ratio']])
    ws3.append(['Budget Accuracy Score', combined_summary.get('budget_accuracy_score', 'N/A')])
    ws3.append([])
    ws3.append(['Ideal Spend-Save-Invest Ratio'])
    for key, val in combined_summary['ideal_vs_actual_ratio']['ideal'].items():
        ws3.append([key.capitalize(), val])
    ws3.append([])
    ws3.append(['Actual Spend-Save-Invest Ratio'])
    for key, val in combined_summary['ideal_vs_actual_ratio']['actual'].items():
        ws3.append([key.capitalize(), val])
    ws3.append([])
    ws3.append(['Category Budget vs Actual'])
    ws3.append(['Category', 'Budget', 'Actual', 'Remaining'])
    for item in combined_summary['budget_vs_actual']:
        ws3.append([item['category'], item['budget'], item['actual'], item['remaining']])
    if 'monthly_cash_flow' in combined_summary:
        ws3.append([])
        ws3.append(['Monthly Cash Flow'])
        ws3.append(['Month', 'Income', 'Expense'])
        for month, values in combined_summary['monthly_cash_flow'].items():
            ws3.append([month, values['income'], values['expense']])
    if 'annual_cash_flow' in combined_summary:
        ws3.append([])
        ws3.append(['Annual Cash Flow'])
        ws3.append(['Year', 'Income', 'Expense'])
        for year, values in combined_summary['annual_cash_flow'].items():
            ws3.append([year, values['income'], values['expense']])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=financial_summary_{period}.xlsx'
    wb.save(response)
    return response
















#setting views


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import UserProfile
from django.contrib import messages

from django.http import JsonResponse

@login_required(login_url='/login')
def setting_view(request):
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)

    investment_options = ["Stocks", "SIPs", "Mutual Funds", "Gold", "Real Estate", "Others"]

    if request.method == 'POST':
        email = request.POST.get('email')
        mobile = request.POST.get('mobileNumber')
        savings = request.POST.get('monthlySavings')
        existing_investments = request.POST.get('existingInvestments')
        risk_appetite = request.POST.get('riskAppetite')
        investment_goals = request.POST.get('investmentGoals')
        preferred_investments = request.POST.getlist('preferredInvestmentType')

        has_changed = (
            user.email != email or
            profile.mobile != mobile or
            profile.savings != savings or
            profile.existing_investments != existing_investments or
            profile.risk_appetite != risk_appetite or
            profile.investment_goals != investment_goals or
            profile.preferred_investments != ", ".join(preferred_investments)
        )

        if has_changed:
            user.email = email
            user.save()

            profile.mobile = mobile
            profile.savings = savings
            profile.existing_investments = existing_investments
            profile.risk_appetite = risk_appetite
            profile.investment_goals = investment_goals
            profile.preferred_investments = ", ".join(preferred_investments)
            profile.save()

            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'message': 'No changes detected'}, status=400)

    selected_options = []
    if profile.preferred_investments:
        selected_options = [opt.strip() for opt in profile.preferred_investments.split(",")]

    context = {
        'profile': profile,
        'investment_options': investment_options,
        'selected_options': selected_options,
        'user_email': user.email,
    }

    return render(request, 'setting.html', context)



from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.http import JsonResponse
from django.shortcuts import render

@login_required(login_url='/login')
def manage_password_view(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important: keeps user logged in
            return JsonResponse({'success': True, 'message': 'Password changed successfully'})
        else:
            errors = form.errors.get_json_data()
            error_messages = [error['message'] for field_errors in errors.values() for error in field_errors]
            return JsonResponse({'success': False, 'errors': error_messages}, status=400)
    else:
        form = PasswordChangeForm(user=request.user)
        return render(request, 'manage_password.html', {'form': form})





from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.http import JsonResponse
import secrets
import string
from .models import DeleteAccountOTP

@login_required(login_url='/login')
def delete_account_view(request):
    return render(request, 'delete_account.html')

@login_required(login_url='/login')
def request_delete_otp(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        confirm_text = request.POST.get('confirm_text')
        password = request.POST.get('password')
        confirm_checkbox = request.POST.get('confirm_checkbox')

        if confirm_text != 'DELETE':
            return JsonResponse({'status': 'error', 'message': 'You must type DELETE exactly to confirm.'})

        if not confirm_checkbox:
            return JsonResponse({'status': 'error', 'message': 'You must confirm by checking the box.'})

        user = authenticate(username=request.user.username, password=password)
        if user is None:
            return JsonResponse({'status': 'error', 'message': 'Password is incorrect.'})

        # Generate OTP using secrets module
        alphabet = string.ascii_uppercase + string.digits
        otp = ''.join(secrets.choice(alphabet) for i in range(6))

        DeleteAccountOTP.objects.filter(user=user).delete() # Delete any existing OTPs
        otp_record = DeleteAccountOTP.objects.create(user=user, otp=otp)

        # Send OTP via email
        subject = 'Confirm Account Deletion'
        message = f"""Dear {user.first_name},

You have requested to delete your Personal Finance Advisor account. To proceed, please enter the following One-Time Password (OTP) on the delete account page: {otp}

This OTP is valid for the next 5 minutes. If you did not request this account deletion, please ignore this email.

Thank you,
The Personal Finance Advisor Team
"""
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [user.email]

        try:
            send_mail(subject, message, from_email, recipient_list, fail_silently=False)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            otp_record.delete() # Delete OTP if email sending fails
            return JsonResponse({'status': 'error', 'message': f'Error sending OTP email: {e}'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})

@login_required(login_url='/login')
def verify_delete_otp(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        otp_entered = request.POST.get('otp')
        user = request.user

        try:
            otp_record = DeleteAccountOTP.objects.get(user=user, otp=otp_entered)
            if (timezone.now() - otp_record.created_at).total_seconds() < 300: # 5 minutes validity
                # OTP is valid, delete the account
                user.delete()
                logout(request)
                return JsonResponse({'status': 'success'})
            else:
                otp_record.delete()
                return JsonResponse({'status': 'error', 'message': 'OTP has expired. Please request again.'})
        except DeleteAccountOTP.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Invalid OTP.'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})









